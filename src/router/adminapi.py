from base64 import decode
from lib2to3.pgen2 import token
import zipfile
from fastapi import APIRouter, Header, Depends, HTTPException, status, Response, UploadFile, Request
from fastapi.responses import FileResponse, StreamingResponse, JSONResponse
from fastapi import FastAPI
from fastapi.routing import APIRoute
from models import query_histroy
from sqlalchemy import null
from sqlmodel import Field, SQLModel, create_engine, Session, select, literal_column, table, desc
from typing import Union
import datetime
import os
import time
import random, string
from pathlib import Path
import sqlite3
from sqlite3 import Connection
from typing import List, Dict
from database import create_db_and_tables, engine
from google.cloud import bigquery
from google.api_core.exceptions import GoogleAPICallError, NotFound
from pydantic import BaseModel
import pandas as pd
import openpyxl
from copy import copy
import json
import logging
from logging.handlers import RotatingFileHandler
import boto3
import io
import tarfile
import gzip
import shutil


DATABASE = 'bq_query.db'

# Logic Apps Workflow URL
LOGIC_APPS_WEBHOOK_URL = "https://prod-02.japaneast.logic.azure.com:443/workflows/38314c48bb6b416a96ced12643cea29c/triggers/manual/paths/invoke?api-version=2016-06-01&sp=%2Ftriggers%2Fmanual%2Frun&sv=1.0&sig=5DJlT6fyzTBMZ1yYi4K75CgmLcDUKrDeFihlifNhUQY"


# ログ設定
logger = logging.getLogger("user_activity")
logger.setLevel(logging.INFO)
handler = RotatingFileHandler(
    "user_activity.log",
    maxBytes=10485760,  # 10MB
    backupCount=10
)
formatter = logging.Formatter('%(message)s')
handler.setFormatter(formatter)
logger.addHandler(handler)

# ログを記録する依存関係関数
async def log_request(request: Request):
    # 現在時刻を取得
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # ユーザーIDを取得（ヘッダーにない場合は "unknown"）
    user_id = request.headers.get("X-User-ID", "unknown")
    
    # エンドポイントのパス
    path = request.url.path
    
    # リクエストパラメータを取得
    params = ""
    
    # GETパラメータを取得
    query_params = dict(request.query_params)
    if query_params:
        params += json.dumps(query_params)
    
    # POSTボディを取得（可能な場合）
    if request.method in ["POST", "PUT"]:
        try:
            # リクエストのボディを複製して内容を確認する
            body_bytes = await request.body()
            
            # ボディがあれば処理する
            if body_bytes:
                try:
                    body = json.loads(body_bytes.decode())
                    if params:
                        params += " "
                    params += json.dumps(body)
                except:
                    if params:
                        params += " "
                    params += "(non-JSON payload)"
            
            # リクエストボディを再度利用可能にする (FastAPIの内部メカニズム)
            request._body = body_bytes
        except:
            if params:
                params += " "
            params += "(body not available)"
    
    # ログ出力 - カンマ区切りのフォーマット
    log_message = f"{timestamp},{user_id},{path},{params}"
    logger.info(log_message)
    
    # 依存関係関数なので何も返さなくてOK
    return

# セッションを取得するための依存関係
def get_session():
    with Session(engine) as session:
        yield session

# router定義 - すべてのエンドポイントにログ依存関係を適用
router = APIRouter(dependencies=[Depends(log_request)])

class SQLQuery(BaseModel):
    sql: str


def send_webhook_notification(title, status, details, success=True):
    """
    Logic Apps Workflowにwebhook通知を送信
    """
    try:
        payload = {
            "title": title,
            "status": status,
            "success": success,
            "timestamp": datetime.datetime.now().isoformat(),
            "details": details,
            "source": "Pallet Cloud Export API"
        }
        
        response = requests.post(
            LOGIC_APPS_WEBHOOK_URL,
            json=payload,
            headers={'Content-Type': 'application/json'},
            timeout=30
        )
        
        print(f"Webhook sent successfully. Status: {response.status_code}")
        return True
        
    except Exception as e:
        print(f"Failed to send webhook: {str(e)}")
        return False
    

@router.post('/gcp/query')
def post_query(query: SQLQuery, session: Session = Depends(get_session)):
    try:
        print(query)
        query_job = client.query(query.sql)  # クエリの実行
        results = query_job.result()  # クエリ結果の取得
        
        # 結果の整形
        rows = []
        for row in results:
            rows.append(dict(row))
        record_count = len(rows)
        session.add(query_histroy(SQL=str(query.sql), last_query_records=record_count))
        session.commit()
        return {"results": rows}
    except (GoogleAPICallError, NotFound) as e:
        raise HTTPException(status_code=400, detail=str(e))


# BigQueryクライアントの初期化
client = bigquery.Client()

# init 処理
@router.on_event("startup")
def on_startup():
    # データベースとテーブルの作成
    create_db_and_tables()

       
@router.get('/', tags=["utils"])
def health_check():
    """
    サーバーヘルスチェック
    """
    return ('200 OK')


@router.get('/queries')
def get_queris(session: Session = Depends(get_session)):
    query = select(query_histroy)
    results = session.exec(query).all()
    return results


@router.post('/query')
def post_query(query_history: query_histroy, session: Session = Depends(get_session)):
    print(query_history)
    print(session)
    session.add(query_history)
    session.commit()
    session.refresh(query_history)
    return query_history


@router.put("/queries/{query_id}")
def update_queries(query_id: int, query_history: query_histroy, session: Session = Depends(get_session)):
    existing_query = session.get(query_histroy, query_id)
    if not existing_query:
        raise HTTPException(status_code=404, detail="Query not found")
    existing_query.title = query_history.title
    existing_query.description = query_history.description
    session.commit()
    session.refresh(existing_query)
    return existing_query
    

@router.post('/login')
def login(item: dict):
    try:
        print(item)
        username = item["username"]
        password = item["password"]
        
        # Original account - full access
        if username == "adi2024" and password == "adi2024":
            return {
                "message": "success",
                "auth": {
                    "rentroll": True,
                    "keiri": True
                },
                "user_id" : 1
            }
        # New account - limited access
        elif username == "adirent2025" and password == "adirent2025":
            return {
                "message": "success",
                "auth": {
                    "rentroll": True,
                    "keiri": False
                },
                "user_id" : 2
            }
        else:
            return {
                "message": "wrong",
                "auth": None
            }
    except:
        return {
            "message": "error",
            "auth": None
        }    

@router.post('/gcp/rentroll')
def post_query(item: dict):
    property_customer_managed_id = item["property_customer_managed_id"]
    date = item["date"]
    try:
        # 賃貸借契約のクエリ
        rentroll_sql = f"""
        SELECT
            *
        FROM
            ard-itandi-production.shared_ard_adi_view.rentroll_output_table as rentroll_output_table
        WHERE
            REGEXP_CONTAINS(rentroll_output_table.property_customer_managed_code, '..{property_customer_managed_id}(-[0-9]+)?')
        ORDER BY
            unit
        """
        
        # 駐車場とバイク置き場のクエリ
        parking_sql = f"""
        SELECT
            *
        FROM
            ard-itandi-production.shared_ard_adi_view.rentroll_parking_output_table as rentroll_parking_output_table
        WHERE
            REGEXP_CONTAINS(rentroll_parking_output_table.property_customer_managed_code, '..{property_customer_managed_id}(-[0-9]+)?')
        ORDER BY
            parking_type, parking_space_number
        """

        # クエリの実行
        rentroll_job = client.query(rentroll_sql)
        parking_job = client.query(parking_sql)

        # 結果の取得
        rentroll_results = rentroll_job.result()
        parking_results = parking_job.result()
        
        # 結果の整形
        rentroll_rows = [dict(row) for row in rentroll_results]
        parking_rows = [dict(row) for row in parking_results]
        
        building_name = rentroll_rows[0]['building_name'] if rentroll_rows else ""
        
        print("rentroll_rows :", rentroll_rows)
        print("parking_rows :", parking_rows)

        # 結果をExcelに整形
        output_filepath = format_to_excel(rentroll_rows, parking_rows, property_customer_managed_id, date, building_name)

        # ファイルを返す
        filename = f"{property_customer_managed_id}_{date}_rentroll.xlsx"
        headers = {"Content-Disposition": f"attachment; filename={filename}"}
        return FileResponse(output_filepath, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers=headers)
        
    except (GoogleAPICallError, NotFound) as e:
       print(e)
       raise HTTPException(status_code=400, detail=str(e))


def write_to_merged_cell(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    if cell.coordinate in ws.merged_cells:
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                top_left = merged_range.min_row, merged_range.min_col
                ws.cell(row=top_left[0], column=top_left[1], value=value)
                return
    ws.cell(row=row, column=col, value=value)


def format_to_excel(rentroll_data, parking_data, property_customer_managed_id, date, building_name):
    # 既存のテンプレートを読み込む
    template_path = './レントロール(原本).xlsx'
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # null値を安全に処理する関数
    def safe_value(value, default=''):
        return value if value is not None else default

    def safe_round(value, decimals=2):
        return round(value, decimals) if value is not None else None

    # データを所定の場所に書き込む
    # title
    write_to_merged_cell(ws, 2, 1, safe_value(building_name))
    write_to_merged_cell(ws, 5, 31, f"{safe_value(date)}時点")

    # 賃貸借契約テーブル 
    start_row = 8
    end_row = 104
    current_row = start_row

    for row in rentroll_data:
        if current_row <= end_row:
            write_to_merged_cell(ws, current_row, 1, safe_value(row.get('floor')))
            write_to_merged_cell(ws, current_row, 2, safe_value(row.get('unit')))
            write_to_merged_cell(ws, current_row, 3, safe_value(row.get('use_type')))
            write_to_merged_cell(ws, current_row, 4, safe_value(row.get('contract_area_m2')))
            write_to_merged_cell(ws, current_row, 5, safe_round(row.get('contract_area_tsubo')))
            write_to_merged_cell(ws, current_row, 6, safe_value(row.get('applicant_name')))
            write_to_merged_cell(ws, current_row, 7, safe_value(row.get('contract_type')))
            write_to_merged_cell(ws, current_row, 8, safe_value(row.get('start_date')))
            write_to_merged_cell(ws, current_row, 9, safe_value(row.get('lease_start_date')))
            write_to_merged_cell(ws, current_row, 10, safe_value(row.get('lease_end_date')))
            write_to_merged_cell(ws, current_row, 11, safe_round(row.get('rent_per_tsubo')))
            write_to_merged_cell(ws, current_row, 12, safe_value(row.get('rent')))
            write_to_merged_cell(ws, current_row, 13, safe_round(row.get('maintenance_fee_per_tsubo')))
            write_to_merged_cell(ws, current_row, 14, safe_value(row.get('maintenance_fee')))
            write_to_merged_cell(ws, current_row, 15, safe_value(row.get('libli_club_monthly_fee')))
            write_to_merged_cell(ws, current_row, 16, safe_value(row.get('tax')))
            write_to_merged_cell(ws, current_row, 17, safe_value(row.get('other_cost')))
            write_to_merged_cell(ws, current_row, 18, safe_value(row.get('other_cost_tax')))
            write_to_merged_cell(ws, current_row, 20, safe_value(row.get('security_deposit_incl_tax')))
            write_to_merged_cell(ws, current_row, 21, safe_value(row.get('key_money_incl_tax')))
            write_to_merged_cell(ws, current_row, 22, safe_value(row.get('guarantee_deposit_incl_tax')))
            write_to_merged_cell(ws, current_row, 23, safe_value(row.get('room_cleaning_fee_upon_move_out_excl_tax')))
            write_to_merged_cell(ws, current_row, 24, safe_value(row.get('cleaning_tax')))
            write_to_merged_cell(ws, current_row, 25, safe_value(row.get('renewal_fee')))
            write_to_merged_cell(ws, current_row, 26, safe_value(row.get('renewal_office_fee')))
            write_to_merged_cell(ws, current_row, 27, safe_value(row.get('renewal_office_fee_tax')))
            write_to_merged_cell(ws, current_row, 28, safe_value(row.get('note')))
            current_row += 1

    # 余った部屋の行を非表示にする
    for row in range(current_row, end_row + 1):
        ws.row_dimensions[row].hidden = True

    # 駐車場の情報を書き込む
    car_parking_start_row = 110
    car_parking_end_row = 151
    car_parking_row = car_parking_start_row

    for row in parking_data:
        if row.get('parking_type') == 'car' and car_parking_row <= car_parking_end_row:
            write_to_merged_cell(ws, car_parking_row, 1, safe_value(row.get('parking_space_number')))
            write_to_merged_cell(ws, car_parking_row, 3, '駐車場')
            write_to_merged_cell(ws, car_parking_row, 6, safe_value(row.get('applicant_name')))
            write_to_merged_cell(ws, car_parking_row, 7, safe_value(row.get('contract_type')))
            write_to_merged_cell(ws, car_parking_row, 8, safe_value(row.get('start_date')))
            write_to_merged_cell(ws, car_parking_row, 9, safe_value(row.get('lease_start_date')))
            write_to_merged_cell(ws, car_parking_row, 10, safe_value(row.get('lease_end_date')))
            parking_fee = safe_value(row.get('parking_fee_excl_tax'), 0)
            write_to_merged_cell(ws, car_parking_row, 11, parking_fee if parking_fee != 0 else None)
            write_to_merged_cell(ws, car_parking_row, 12, safe_value(row.get('parking_fee_tax')))
            write_to_merged_cell(ws, car_parking_row, 14, safe_value(row.get('security_deposit_incl_tax')))
            write_to_merged_cell(ws, car_parking_row, 15, safe_value(row.get('key_money_incl_tax')))
            write_to_merged_cell(ws, car_parking_row, 16, safe_value(row.get('renewal_fee')))
            write_to_merged_cell(ws, car_parking_row, 17, safe_value(row.get('renewal_office_fee')))
            write_to_merged_cell(ws, car_parking_row, 18, safe_value(row.get('renewal_office_fee_tax')))
            car_parking_row += 1

    # 余った駐車場の行を非表示にする
    for row in range(car_parking_row, car_parking_end_row + 1):
        ws.row_dimensions[row].hidden = True

    # バイク置き場の情報を書き込む
    motorbike_parking_start_row = 157
    motorbike_parking_end_row = 196
    motorbike_parking_row = motorbike_parking_start_row

    for row in parking_data:
        if row.get('parking_type') == 'motorbike' and motorbike_parking_row <= motorbike_parking_end_row:
            write_to_merged_cell(ws, motorbike_parking_row, 1, safe_value(row.get('parking_space_number')))
            write_to_merged_cell(ws, motorbike_parking_row, 3, 'バイク置き場')
            write_to_merged_cell(ws, motorbike_parking_row, 6, safe_value(row.get('applicant_name')))
            write_to_merged_cell(ws, motorbike_parking_row, 7, safe_value(row.get('contract_type')))
            write_to_merged_cell(ws, motorbike_parking_row, 8, safe_value(row.get('start_date')))
            write_to_merged_cell(ws, motorbike_parking_row, 9, safe_value(row.get('lease_start_date')))
            write_to_merged_cell(ws, motorbike_parking_row, 10, safe_value(row.get('lease_end_date')))
            motorcycle_fee = safe_value(row.get('motorcycle_parking_fee_excl_tax'), 0)
            write_to_merged_cell(ws, motorbike_parking_row, 11, motorcycle_fee if motorcycle_fee != 0 else None)
            write_to_merged_cell(ws, motorbike_parking_row, 12, safe_value(row.get('motorcycle_parking_fee_tax')))
            write_to_merged_cell(ws, motorbike_parking_row, 14, safe_value(row.get('security_deposit_incl_tax')))
            write_to_merged_cell(ws, motorbike_parking_row, 15, safe_value(row.get('key_money_incl_tax')))
            write_to_merged_cell(ws, motorbike_parking_row, 16, safe_value(row.get('renewal_fee')))
            write_to_merged_cell(ws, motorbike_parking_row, 17, safe_value(row.get('renewal_office_fee')))
            write_to_merged_cell(ws, motorbike_parking_row, 18, safe_value(row.get('renewal_office_fee_tax')))
            motorbike_parking_row += 1

    # 余ったバイク置き場の行を非表示にする
    for row in range(motorbike_parking_row, motorbike_parking_end_row + 1):
        ws.row_dimensions[row].hidden = True

    # 出力ファイル名
    output_filename = f"{property_customer_managed_id}_{date}_rentroll.xlsx"
    output_filepath = f"./rentroll/{output_filename}"
    
    # ファイルに保存
    wb.save(output_filepath)
    
    return output_filepath

# パラメータを受け取るためのモデルを追加
class SuitotyoParams(BaseModel):
    start_date: str
    end_date: str
    account: str

class HosyoKaisyaParams(BaseModel):
   account_year: int 
   account_month: int
   
@router.post('/gcp/suitotyo')
def get_suitotyo(params: SuitotyoParams):
    try:
        # 出納帳のクエリ
        suitotyo_sql = f"""
        SELECT
            *
        FROM
            `ard-itandi-production.shared_ard_adi_view.suitotyo`
        WHERE
            `入金日` between "{params.start_date}" and "{params.end_date}"
            AND `口座` = "{params.account}"
        """

        print(suitotyo_sql)
        
        # クエリの実行
        query_job = client.query(suitotyo_sql)
        results = query_job.result()
        
        # 結果の整形
        rows = [dict(row) for row in results]
        
        return {"results": rows}
        
    except (GoogleAPICallError, NotFound) as e:
        print(e)
        raise HTTPException(status_code=400, detail=str(e))


@router.post('/gcp/hosyo-kaisya-unmatch')
def get_hosyo_kaisya_unmatch(params: HosyoKaisyaParams):
   try:
       hosyo_sql = f"""
       SELECT
           *
       FROM
           `ard-itandi-production.shared_ard_adi_view.hosyo_kaisya_matchlist` 
       WHERE
           account_year = {params.account_year}
           AND account_month = {params.account_month}
       """

       print(hosyo_sql)
       
       query_job = client.query(hosyo_sql)
       results = query_job.result()

       rows = [dict(row) for row in results]
       return {"results": rows}
       
   except (GoogleAPICallError, NotFound) as e:
       print(e)
       raise HTTPException(status_code=400, detail=str(e))

@router.post('/gcp/pallet-cloud')
def export_to_pallet_cloud():
    today = datetime.datetime.now().strftime("%Y%m%d")
    
    try:
        # 出力先のS3バケットとプレフィックス
        bucket_name = "adi-external-integration"
        prefix = "pallet-cloud/prod/"
        
        # S3クライアントの初期化
        s3_client = boto3.client('s3')
        
        # エクスポートするテーブルとファイル名のマッピング
        table_file_mapping = {
            "PC_buildings_output": f"{today}_PC_buildings",
            "PC_rooms_output": f"{today}_PC_rooms",
            "PC_tenants_output": f"{today}_PC_tenants",
            "PC_contract2": f"{today}_PC_contract2",
            "PC_contract_tenant": f"{today}_PC_contract_tenant",
            "PC_contract_resident": f"{today}_PC_contract_resident",
            "PC_commitment": f"{today}_PC_commitment"
        }
        
        results = {}
        total_rows = 0
        
        # 各テーブルのデータをクエリしてS3にエクスポート
        for table_name, file_name in table_file_mapping.items():
            try:
                print(f"Processing table: {table_name}")
                
                # BigQueryからデータを取得
                query = f"""
                SELECT *
                FROM `ard-itandi-production.shared_ard_adi_view.{table_name}`
                """
                
                query_job = client.query(query)
                df = query_job.to_dataframe()
                
                # DataFrameをCSVに変換 (UTF-8エンコーディングを指定)
                csv_buffer = io.StringIO()
                df.to_csv(csv_buffer, index=False)
                
                # S3にアップロード (UTF-8エンコーディングを明示的に指定)
                s3_key = f"{prefix}{file_name}.csv"
                s3_client.put_object(
                    Bucket=bucket_name,
                    Key=s3_key,
                    Body=csv_buffer.getvalue().encode('utf-8'),
                    ContentType='text/csv; charset=utf-8'
                )
                
                row_count = len(df)
                total_rows += row_count
                
                results[table_name] = {
                    "status": "success",
                    "rows": row_count,
                    "destination": f"s3://{bucket_name}/{s3_key}"
                }
                
                print(f"Successfully exported {table_name}: {row_count} rows")
                
            except Exception as table_error:
                print(f"Error processing table {table_name}: {str(table_error)}")
                results[table_name] = {
                    "status": "error",
                    "error": str(table_error),
                    "rows": 0
                }
        
        # 成功時のwebhook通知を送信
        success_details = {
            "export_date": today,
            "total_files": len(table_file_mapping),
            "total_rows": total_rows,
            "file_details": [
                {
                    "table": table_name,
                    "filename": f"{table_file_mapping[table_name]}.csv",
                    "rows": result.get("rows", 0),
                    "status": result.get("status", "unknown")
                }
                for table_name, result in results.items()
            ],
            "s3_bucket": bucket_name,
            "s3_prefix": prefix
        }
        
        # エラーがあったかチェック
        failed_tables = [table for table, result in results.items() if result.get("status") == "error"]
        
        if failed_tables:
            # 一部失敗した場合
            webhook_title = "パレットクラウド向けファイル出力 - 一部エラー"
            webhook_status = f"完了（{len(failed_tables)}件のエラーあり）"
            success_details["failed_tables"] = failed_tables
            send_webhook_notification(webhook_title, webhook_status, success_details, success=False)
        else:
            # 全て成功した場合
            webhook_title = "パレットクラウド向けファイル出力 - 完了"
            webhook_status = f"全{len(table_file_mapping)}ファイル正常出力（計{total_rows:,}件）"
            send_webhook_notification(webhook_title, webhook_status, success_details, success=True)
        
        return {
            "message": "データのエクスポートが完了しました",
            "date": today,
            "total_rows": total_rows,
            "results": results
        }
        
    except (GoogleAPICallError, NotFound) as e:
        error_message = f"BigQuery Error: {str(e)}"
        print(error_message)
        
        # BigQueryエラー時のwebhook通知
        error_details = {
            "export_date": today,
            "error_type": "BigQuery Error",
            "error_message": str(e),
            "tables_attempted": list(table_file_mapping.keys())
        }
        
        send_webhook_notification(
            "パレットクラウド向けファイル出力 - BigQueryエラー",
            "BigQueryアクセスに失敗しました",
            error_details,
            success=False
        )
        
        raise HTTPException(status_code=400, detail=str(e))
        
    except Exception as e:
        error_message = f"Unexpected Error: {str(e)}"
        print(error_message)
        
        # 予期しないエラー時のwebhook通知
        error_details = {
            "export_date": today,
            "error_type": "Unexpected Error",
            "error_message": str(e),
            "tables_attempted": list(table_file_mapping.keys())
        }
        
        send_webhook_notification(
            "パレットクラウド向けファイル出力 - システムエラー",
            "予期しないエラーが発生しました",
            error_details,
            success=False
        )
        
        raise HTTPException(status_code=500, detail=str(e))

class DaysAgoParams(BaseModel):
    days_ago: int = 0

@router.post('/gcp/pallet-cloud/rooms-diff')
def rooms_diff(params: DaysAgoParams = None):
    try:
        # パラメータがない場合はデフォルト値を使用
        days_ago = 0 if params is None else params.days_ago
        
        # 現在の日付と指定された日数前の日付をyyyymmdd形式で取得
        today = datetime.datetime.now()
        base_date = today - datetime.timedelta(days=days_ago)
        yesterday = base_date - datetime.timedelta(days=1)
        today_str = base_date.strftime("%Y%m%d")
        yesterday_str = yesterday.strftime("%Y%m%d")
        
        # S3バケットとプレフィックス
        bucket_name = "adi-external-integration"
        prefix = "pallet-cloud/prod/"
        
        # 今日と昨日のファイルパス
        today_file = f"{today_str}_PC_rooms.csv"
        yesterday_file = f"{yesterday_str}_PC_rooms.csv"
        
        today_s3_path = f"{prefix}{today_file}"
        yesterday_s3_path = f"{prefix}{yesterday_file}"
        
        # S3クライアントの初期化
        s3_client = boto3.client('s3')
        
        # 一時ファイルパス
        temp_today_file = f"/tmp/{today_file}"
        temp_yesterday_file = f"/tmp/{yesterday_file}"
        
        # S3からファイルをダウンロード
        try:
            s3_client.download_file(bucket_name, today_s3_path, temp_today_file)
            print(f"Downloaded today's file from s3://{bucket_name}/{today_s3_path}")
        except Exception as e:
            print(f"Error downloading today's file: {e}")
            raise HTTPException(status_code=404, detail=f"Today's file not found: {today_file}")
            
        try:
            s3_client.download_file(bucket_name, yesterday_s3_path, temp_yesterday_file)
            print(f"Downloaded yesterday's file from s3://{bucket_name}/{yesterday_s3_path}")
        except Exception as e:
            print(f"Error downloading yesterday's file: {e}")
            raise HTTPException(status_code=404, detail=f"Yesterday's file not found: {yesterday_file}")
        
        # CSVファイルをDataFrameに読み込む（データクリーニング付き）
        def clean_dataframe(df):
            """DataFrameの文字列カラムをクリーニング"""
            for col in df.columns:
                if df[col].dtype == 'object':  # 文字列カラムの場合
                    # 空文字列をNaNに統一
                    df[col] = df[col].replace('', pd.NA)
                    # 先頭・末尾の空白を除去
                    df[col] = df[col].astype(str).str.strip()
                    # 'nan'文字列をNaNに変換
                    df[col] = df[col].replace('nan', pd.NA)
                    # 空白のみの文字列をNaNに変換
                    df[col] = df[col].replace(r'^\s*$', pd.NA, regex=True)
            return df
        
        today_df = pd.read_csv(temp_today_file, encoding='utf-8', keep_default_na=True, na_values=['', 'NULL', 'null', 'None'])
        yesterday_df = pd.read_csv(temp_yesterday_file, encoding='utf-8', keep_default_na=True, na_values=['', 'NULL', 'null', 'None'])
        
        # データクリーニング
        today_df = clean_dataframe(today_df)
        yesterday_df = clean_dataframe(yesterday_df)
        
        print(f"Today's file has {len(today_df)} rows")
        print(f"Yesterday's file has {len(yesterday_df)} rows")
        
        # 元のカラム順序を保存
        original_columns = today_df.columns.tolist()
        
        # 主キーを特定（32番目のカラムを主キーとする）
        if len(today_df.columns) > 31:  # 0-indexedなので31が32番目
            primary_key = today_df.columns[31]
            
            # **最適化1: インデックスを設定して高速化**
            today_df_indexed = today_df.set_index(primary_key)
            yesterday_df_indexed = yesterday_df.set_index(primary_key)
            
            # **最適化2: 集合演算を使用**
            today_keys = set(today_df_indexed.index)
            yesterday_keys = set(yesterday_df_indexed.index)
            
            # 1. 今日のファイルにしかない行を抽出
            only_today_keys = today_keys - yesterday_keys
            only_in_today = today_df_indexed.loc[list(only_today_keys)] if only_today_keys else pd.DataFrame()
            
            # 2. 共通キーを取得
            common_keys = today_keys & yesterday_keys
            print(f"Processing {len(common_keys)} common keys for differences")
            
            # **最適化3: ベクトル化された比較**
            diff_details = []
            changed_indices = set()  # setを使用して重複排除を効率化
            detailed_diff_info = []  # デバッグ用の詳細情報
            
            if common_keys:
                # 共通キーのデータを一括で取得
                today_common = today_df_indexed.loc[list(common_keys)]
                yesterday_common = yesterday_df_indexed.loc[list(common_keys)]
                
                # 比較から除外するカラム（インデックス）
                exclude_indices = [19]
                include_columns = [col for i, col in enumerate(today_common.columns) if i not in exclude_indices]
                
                # **最適化4: カラムごとに一括比較（改良版）**
                for col in include_columns:
                    if col in today_common.columns and col in yesterday_common.columns:
                        today_vals = today_common[col]
                        yesterday_vals = yesterday_common[col]
                        
                        # 数値型の場合は差分が1以上あるかチェック
                        if pd.api.types.is_numeric_dtype(today_vals):
                            # 両方がNaNでない場合の数値差分チェック
                            both_not_nan = pd.notna(today_vals) & pd.notna(yesterday_vals)
                            diff_mask = both_not_nan & (abs(today_vals - yesterday_vals) >= 1)
                            # 片方だけがNaNの場合もチェック
                            nan_diff_mask = today_vals.isna() != yesterday_vals.isna()
                            combined_mask = diff_mask | nan_diff_mask
                        else:
                            # 文字列型の場合：改良された比較ロジック
                            # NaNの扱いを統一（両方がNaNの場合は差分なし）
                            both_nan = pd.isna(today_vals) & pd.isna(yesterday_vals)
                            both_not_nan = pd.notna(today_vals) & pd.notna(yesterday_vals)
                            one_nan = pd.isna(today_vals) != pd.isna(yesterday_vals)
                            
                            # 両方が値を持つ場合の比較
                            value_diff = both_not_nan & (today_vals.astype(str) != yesterday_vals.astype(str))
                            
                            # 差分があるのは：値が異なる場合 OR 片方だけがNaNの場合
                            combined_mask = value_diff | one_nan
                        
                        # 差分があるキーをsetに追加（自動重複排除）
                        different_keys = combined_mask[combined_mask].index.tolist()
                        if different_keys:
                            changed_indices.update(different_keys)
                            
                            # デバッグ用：差分がある全件の詳細情報を記録
                            for key in different_keys:
                                today_val = today_vals.loc[key] if key in today_vals.index else 'KEY_NOT_FOUND'
                                yesterday_val = yesterday_vals.loc[key] if key in yesterday_vals.index else 'KEY_NOT_FOUND'
                                detailed_diff_info.append({
                                    'key': key,
                                    'column': col,
                                    'today_value': repr(today_val),
                                    'yesterday_value': repr(yesterday_val),
                                    'today_type': type(today_val).__name__,
                                    'yesterday_type': type(yesterday_val).__name__
                                })
                
                # setをリストに変換
                changed_indices = list(changed_indices)
                
                # 差分詳細記録
                for key in list(changed_indices):
                    diff_details.append({"key": key, "differences": ["Changes detected"]})
            
            # 変更された行を取得
            changed_rows = today_df_indexed.loc[changed_indices] if changed_indices else pd.DataFrame()
            
            # 差分ファイルを作成（今日にしかない行 + 変更された行）
            diff_frames = []
            if not only_in_today.empty:
                # インデックスをリセットして元のカラム順序を復元
                only_in_today_reset = only_in_today.reset_index()
                only_in_today_ordered = only_in_today_reset[original_columns]
                diff_frames.append(only_in_today_ordered)
            
            if not changed_rows.empty:
                # インデックスをリセットして元のカラム順序を復元
                changed_rows_reset = changed_rows.reset_index()
                changed_rows_ordered = changed_rows_reset[original_columns]
                diff_frames.append(changed_rows_ordered)
            
            if diff_frames:
                diff_df = pd.concat(diff_frames, ignore_index=True)
            else:
                diff_df = pd.DataFrame()
            
            # 差分ファイルをCSVに変換
            csv_buffer = io.StringIO()
            diff_df.to_csv(csv_buffer, index=False)
            
            # S3に出力（UTF-8エンコーディングを明示的に指定）
            output_s3_key = f"{prefix}output/room.csv"
            s3_client.put_object(
                Bucket=bucket_name,
                Key=output_s3_key,
                Body=csv_buffer.getvalue().encode('utf-8'),
                ContentType='text/csv; charset=utf-8'
            )
            
            # ローカルにも保存
            local_dir = f"./pallet_cloud/{today_str}"
            os.makedirs(local_dir, exist_ok=True)
            local_file_path = f"{local_dir}/room.csv"
            
            # CSVファイルを保存
            with open(local_file_path, 'w', encoding='utf-8') as f:
                f.write(csv_buffer.getvalue())
            
            # 差分詳細をテキストファイルに出力（デバッグ情報付き）
            diff_result_path = f"{local_dir}/room-diff-result.txt"
            with open(diff_result_path, 'w', encoding='utf-8') as f:
                f.write(f"Room Diff Results - {today_str} vs {yesterday_str}\n")
                f.write(f"Total rows in today's file: {len(today_df)}\n")
                f.write(f"Total rows in yesterday's file: {len(yesterday_df)}\n")
                f.write(f"New rows (only in today's file): {len(only_in_today)}\n")
                f.write(f"Changed rows: {len(changed_rows)}\n\n")
                
                if len(only_in_today) > 0:
                    f.write("=== NEW ROWS ===\n")
                    for key in only_in_today.index:
                        f.write(f"New row with key: {key}\n")
                    f.write("\n")
                
                if len(diff_details) > 0:
                    f.write("=== CHANGED ROWS ===\n")
                    for row_diff in diff_details:
                        f.write(f"Row with key '{row_diff['key']}' has differences\n")
                    f.write("\n")
                    
                    # デバッグ用の詳細情報を出力
                    if detailed_diff_info:
                        f.write(f"=== DETAILED DIFF INFO (All {len(detailed_diff_info)} differences) ===\n")
                        for info in detailed_diff_info:
                            f.write(f"Key: {info['key']}, Column: {info['column']}\n")
                            f.write(f"  Today: {info['today_value']} (type: {info['today_type']})\n")
                            f.write(f"  Yesterday: {info['yesterday_value']} (type: {info['yesterday_type']})\n")
                            f.write("\n")
                else:
                    f.write("No differences found in existing rows.\n")
            
            # 差分詳細ファイルをS3にもアップロード
            with open(diff_result_path, 'r', encoding='utf-8') as f:
                diff_detail_content = f.read()
            
            # S3に出力
            diff_detail_s3_key = f"{prefix}output/room-diff-result.txt"
            s3_client.put_object(
                Bucket=bucket_name,
                Key=diff_detail_s3_key,
                Body=diff_detail_content.encode('utf-8'),
                ContentType='text/plain; charset=utf-8'
            )
            
            print(f"Saved local file to: {local_file_path}")
            print(f"Saved diff details to: {diff_result_path}")
            print(f"Uploaded diff details to S3: s3://{bucket_name}/{diff_detail_s3_key}")
            
            output_s3_path = f"s3://{bucket_name}/{output_s3_key}"
            
            return {
                "message": "部屋ファイルの差分計算が完了しました",
                "today_date": today_str,
                "yesterday_date": yesterday_str,
                "total_rows_today": len(today_df),
                "total_rows_yesterday": len(yesterday_df),
                "new_rows": len(only_in_today),
                "changed_rows": len(changed_indices),
                "diff_rows": len(diff_df),
                "output_file": output_s3_path,
                "debug_info": {
                    "detailed_diffs_found": len(detailed_diff_info),
                    "sample_differences": detailed_diff_info[:3] if detailed_diff_info else [],
                    "excluded_columns": "Column at index 19 excluded from comparison"
                }
            }
        else:
            raise HTTPException(status_code=400, detail="ファイルに32番目のカラムがありません")
            
    except (GoogleAPICallError, NotFound) as e:
        print(f"BigQuery Error: {e}")
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        print(f"Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))    

@router.post('/gcp/pallet-cloud/contract2-diff')
def contract2_diff(params: DaysAgoParams = None):
    try:
        # パラメータがない場合はデフォルト値を使用
        days_ago = 0 if params is None else params.days_ago
        
        # 現在の日付と指定された日数前の日付をyyyymmdd形式で取得
        today = datetime.datetime.now()
        base_date = today - datetime.timedelta(days=days_ago)
        yesterday = base_date - datetime.timedelta(days=1)
        today_str = base_date.strftime("%Y%m%d")
        yesterday_str = yesterday.strftime("%Y%m%d")
        
        # S3バケットとプレフィックス
        bucket_name = "adi-external-integration"
        prefix = "pallet-cloud/prod/"
        
        # 今日と昨日のファイルパス
        today_file = f"{today_str}_PC_contract2.csv"
        yesterday_file = f"{yesterday_str}_PC_contract2.csv"
        
        today_s3_path = f"{prefix}{today_file}"
        yesterday_s3_path = f"{prefix}{yesterday_file}"
        
        # S3クライアントの初期化
        s3_client = boto3.client('s3')
        
        # 一時ファイルパス
        temp_today_file = f"/tmp/{today_file}"
        temp_yesterday_file = f"/tmp/{yesterday_file}"
        
        # S3からファイルをダウンロード
        try:
            s3_client.download_file(bucket_name, today_s3_path, temp_today_file)
            print(f"Downloaded today's file from s3://{bucket_name}/{today_s3_path}")
        except Exception as e:
            print(f"Error downloading today's file: {e}")
            raise HTTPException(status_code=404, detail=f"Today's file not found: {today_file}")
            
        try:
            s3_client.download_file(bucket_name, yesterday_s3_path, temp_yesterday_file)
            print(f"Downloaded yesterday's file from s3://{bucket_name}/{yesterday_s3_path}")
        except Exception as e:
            print(f"Error downloading yesterday's file: {e}")
            raise HTTPException(status_code=404, detail=f"Yesterday's file not found: {yesterday_file}")
        
        # CSVファイルをDataFrameに読み込む（データクリーニング付き）
        def clean_dataframe(df):
            """DataFrameの文字列カラムをクリーニング"""
            for col in df.columns:
                if df[col].dtype == 'object':  # 文字列カラムの場合
                    # 空文字列をNaNに統一
                    df[col] = df[col].replace('', pd.NA)
                    # 先頭・末尾の空白を除去
                    df[col] = df[col].astype(str).str.strip()
                    # 'nan'文字列をNaNに変換
                    df[col] = df[col].replace('nan', pd.NA)
                    # 空白のみの文字列をNaNに変換
                    df[col] = df[col].replace(r'^\s*$', pd.NA, regex=True)
            return df
        
        today_df = pd.read_csv(temp_today_file, encoding='utf-8', keep_default_na=True, na_values=['', 'NULL', 'null', 'None'])
        yesterday_df = pd.read_csv(temp_yesterday_file, encoding='utf-8', keep_default_na=True, na_values=['', 'NULL', 'null', 'None'])
        
        # データクリーニング前の情報をデバッグ出力
        print("=== BEFORE CLEANING DEBUG INFO ===")
        print(f"Today original dtypes:\n{today_df.dtypes}")
        print(f"Yesterday original dtypes:\n{yesterday_df.dtypes}")
        
        # データクリーニング
        today_df = clean_dataframe(today_df)
        yesterday_df = clean_dataframe(yesterday_df)
        
        # データクリーニング後の情報をデバッグ出力
        print("=== AFTER CLEANING DEBUG INFO ===")
        print(f"Today after cleaning dtypes:\n{today_df.dtypes}")
        print(f"Yesterday after cleaning dtypes:\n{yesterday_df.dtypes}")
        
        print(f"Today's file has {len(today_df)} rows")
        print(f"Yesterday's file has {len(yesterday_df)} rows")
        
        # 元のカラム順序を保存
        original_columns = today_df.columns.tolist()
        
        # 主キーを特定（1列目のカラムを主キーとする）
        if len(today_df.columns) > 0:
            primary_key = today_df.columns[0]
            print(f"Primary key: {primary_key}")
            
            # 主キーの重複をチェック
            today_duplicates = today_df[primary_key].duplicated().sum()
            yesterday_duplicates = yesterday_df[primary_key].duplicated().sum()
            print(f"Duplicates check - Today: {today_duplicates}, Yesterday: {yesterday_duplicates}")
            
            if today_duplicates > 0 or yesterday_duplicates > 0:
                print(f"Warning: Found duplicates in primary key. Today: {today_duplicates}, Yesterday: {yesterday_duplicates}")
                # 重複を削除
                today_df = today_df.drop_duplicates(subset=[primary_key], keep='first')
                yesterday_df = yesterday_df.drop_duplicates(subset=[primary_key], keep='first')
                print(f"After removing duplicates - Today: {len(today_df)} rows, Yesterday: {len(yesterday_df)} rows")
            
            # **最適化1: インデックスを設定して高速化**
            try:
                today_df_indexed = today_df.set_index(primary_key)
                yesterday_df_indexed = yesterday_df.set_index(primary_key)
                print("Successfully set index for both dataframes")
            except Exception as e:
                print(f"ERROR setting index: {e}")
                raise e
            
            # **最適化2: 集合演算を使用**
            today_keys = set(today_df_indexed.index)
            yesterday_keys = set(yesterday_df_indexed.index)
            
            # 1. 今日のファイルにしかない行を抽出
            only_today_keys = today_keys - yesterday_keys
            only_in_today = today_df_indexed.loc[list(only_today_keys)] if only_today_keys else pd.DataFrame()
            
            # 2. 共通キーを取得
            common_keys = today_keys & yesterday_keys
            print(f"Processing {len(common_keys)} common keys for differences")
            
            # **最適化3: ベクトル化された比較**
            diff_details = []
            changed_indices = set()  # setを使用して重複排除を効率化
            detailed_diff_info = []  # デバッグ用の詳細情報
            
            if common_keys:
                # 共通キーのデータを一括で取得
                try:
                    today_common = today_df_indexed.loc[list(common_keys)]
                    yesterday_common = yesterday_df_indexed.loc[list(common_keys)]
                    print("Successfully extracted common keys data")
                except Exception as e:
                    print(f"ERROR extracting common keys data: {e}")
                    raise e
                
                # デバッグ情報を出力
                print("=== COMPARISON DEBUG INFO ===")
                print(f"Today common shape: {today_common.shape}")
                print(f"Yesterday common shape: {yesterday_common.shape}")
                print(f"Today common columns: {today_common.columns.tolist()}")
                print(f"Yesterday common columns: {yesterday_common.columns.tolist()}")
                print(f"Today common dtypes:\n{today_common.dtypes}")
                print(f"Yesterday common dtypes:\n{yesterday_common.dtypes}")
                
                # カラム名の一致確認
                if not today_common.columns.equals(yesterday_common.columns):
                    print("ERROR: Column names do not match!")
                    print(f"Today only: {set(today_common.columns) - set(yesterday_common.columns)}")
                    print(f"Yesterday only: {set(yesterday_common.columns) - set(today_common.columns)}")
                    raise Exception("Column names do not match between today and yesterday data")
                
                # 比較から除外するカラム（インデックス）
                # 元々20番目だったが、インデックス設定により20番目のカラムを除外
                exclude_indices = [20]  
                include_columns = [col for i, col in enumerate(today_common.columns) if i not in exclude_indices]
                print(f"Include columns for comparison: {include_columns}")
                
                # **最適化4: カラムごとに一括比較（改良版）**
                for col_index, col in enumerate(include_columns):
                    try:
                        print(f"Processing column {col_index + 1}/{len(include_columns)}: '{col}'")
                        
                        if col in today_common.columns and col in yesterday_common.columns:
                            today_vals = today_common[col]
                            yesterday_vals = yesterday_common[col]
                            
                            print(f"  Column '{col}' - Today dtype: {today_vals.dtype}, Yesterday dtype: {yesterday_vals.dtype}")
                            
                            # データ型の一致確認
                            if today_vals.dtype != yesterday_vals.dtype:
                                print(f"  WARNING: dtype mismatch for column '{col}' - Today: {today_vals.dtype}, Yesterday: {yesterday_vals.dtype}")
                            
                            # 数値型の場合は差分が1以上あるかチェック
                            if pd.api.types.is_numeric_dtype(today_vals):
                                print(f"  Processing as numeric column")
                                # 両方がNaNでない場合の数値差分チェック
                                both_not_nan = pd.notna(today_vals) & pd.notna(yesterday_vals)
                                diff_mask = both_not_nan & (abs(today_vals - yesterday_vals) >= 1)
                                # 片方だけがNaNの場合もチェック
                                nan_diff_mask = today_vals.isna() != yesterday_vals.isna()
                                combined_mask = diff_mask | nan_diff_mask
                            else:
                                print(f"  Processing as string column")
                                # 文字列型の場合：改良された比較ロジック
                                # NaNの扱いを統一（両方がNaNの場合は差分なし）
                                both_nan = pd.isna(today_vals) & pd.isna(yesterday_vals)
                                both_not_nan = pd.notna(today_vals) & pd.notna(yesterday_vals)
                                one_nan = pd.isna(today_vals) != pd.isna(yesterday_vals)
                                
                                # 両方が値を持つ場合の比較
                                try:
                                    value_diff = both_not_nan & (today_vals.astype(str) != yesterday_vals.astype(str))
                                except Exception as e:
                                    print(f"  ERROR in string comparison for column '{col}': {e}")
                                    print(f"  Today sample values: {today_vals.head()}")
                                    print(f"  Yesterday sample values: {yesterday_vals.head()}")
                                    raise e
                                
                                # 差分があるのは：値が異なる場合 OR 片方だけがNaNの場合
                                combined_mask = value_diff | one_nan
                            
                            # 差分があるキーをsetに追加（自動重複排除）
                            different_keys = combined_mask[combined_mask].index.tolist()
                            if different_keys:
                                changed_indices.update(different_keys)
                                print(f"  Found {len(different_keys)} differences in column '{col}'")
                                
                                # デバッグ用：差分がある全件の詳細情報を記録
                                for key in different_keys:
                                    today_val = today_vals.loc[key] if key in today_vals.index else 'KEY_NOT_FOUND'
                                    yesterday_val = yesterday_vals.loc[key] if key in yesterday_vals.index else 'KEY_NOT_FOUND'
                                    detailed_diff_info.append({
                                        'key': key,
                                        'column': col,
                                        'today_value': repr(today_val),
                                        'yesterday_value': repr(yesterday_val),
                                        'today_type': type(today_val).__name__,
                                        'yesterday_type': type(yesterday_val).__name__
                                    })
                            else:
                                print(f"  No differences found in column '{col}'")
                        
                    except Exception as e:
                        print(f"ERROR processing column '{col}': {e}")
                        print(f"Column index: {col_index}")
                        if col in today_common.columns:
                            print(f"Today column sample: {today_common[col].head()}")
                        if col in yesterday_common.columns:
                            print(f"Yesterday column sample: {yesterday_common[col].head()}")
                        raise e
                
                # setをリストに変換
                changed_indices = list(changed_indices)
                print(f"Total changed indices: {len(changed_indices)}")
                
                # 差分詳細記録
                for key in list(changed_indices):
                    diff_details.append({"key": key, "differences": ["Changes detected"]})
            
            # 変更された行を取得
            changed_rows = today_df_indexed.loc[changed_indices] if changed_indices else pd.DataFrame()
            
            # 差分ファイルを作成（今日にしかない行 + 変更された行）
            diff_frames = []
            if not only_in_today.empty:
                # インデックスをリセットして元のカラム順序を復元
                only_in_today_reset = only_in_today.reset_index()
                only_in_today_ordered = only_in_today_reset[original_columns]
                diff_frames.append(only_in_today_ordered)
            
            if not changed_rows.empty:
                # インデックスをリセットして元のカラム順序を復元
                changed_rows_reset = changed_rows.reset_index()
                changed_rows_ordered = changed_rows_reset[original_columns]
                diff_frames.append(changed_rows_ordered)
            
            if diff_frames:
                diff_df = pd.concat(diff_frames, ignore_index=True)
            else:
                diff_df = pd.DataFrame()
            
            # 差分ファイルをCSVに変換
            csv_buffer = io.StringIO()
            diff_df.to_csv(csv_buffer, index=False)
            
            # S3に出力（UTF-8エンコーディングを明示的に指定）
            output_s3_key = f"{prefix}output/contract2.csv"
            s3_client.put_object(
                Bucket=bucket_name,
                Key=output_s3_key,
                Body=csv_buffer.getvalue().encode('utf-8'),
                ContentType='text/csv; charset=utf-8'
            )
            
            # ローカルにも保存
            local_dir = f"./pallet_cloud/{today_str}"
            os.makedirs(local_dir, exist_ok=True)
            local_file_path = f"{local_dir}/contract2.csv"
            
            # CSVファイルを保存
            with open(local_file_path, 'w', encoding='utf-8') as f:
                f.write(csv_buffer.getvalue())
            
            # 差分詳細をテキストファイルに出力（デバッグ情報付き）
            diff_result_path = f"{local_dir}/contract2-diff-result.txt"
            with open(diff_result_path, 'w', encoding='utf-8') as f:
                f.write(f"Contract2 Diff Results - {today_str} vs {yesterday_str}\n")
                f.write(f"Total rows in today's file: {len(today_df)}\n")
                f.write(f"Total rows in yesterday's file: {len(yesterday_df)}\n")
                f.write(f"New rows (only in today's file): {len(only_in_today)}\n")
                f.write(f"Changed rows: {len(changed_rows)}\n\n")
                
                if len(only_in_today) > 0:
                    f.write("=== NEW ROWS ===\n")
                    for key in only_in_today.index:
                        f.write(f"New row with key: {key}\n")
                    f.write("\n")
                
                if len(diff_details) > 0:
                    f.write("=== CHANGED ROWS ===\n")
                    for row_diff in diff_details:
                        f.write(f"Row with key '{row_diff['key']}' has differences\n")
                    f.write("\n")
                    
                    # デバッグ用の詳細情報を出力
                    if detailed_diff_info:
                        f.write(f"=== DETAILED DIFF INFO (All {len(detailed_diff_info)} differences) ===\n")
                        for info in detailed_diff_info:
                            f.write(f"Key: {info['key']}, Column: {info['column']}\n")
                            f.write(f"  Today: {info['today_value']} (type: {info['today_type']})\n")
                            f.write(f"  Yesterday: {info['yesterday_value']} (type: {info['yesterday_type']})\n")
                            f.write("\n")
                else:
                    f.write("No differences found in existing rows.\n")
            
            # 差分詳細ファイルをS3にもアップロード
            with open(diff_result_path, 'r', encoding='utf-8') as f:
                diff_detail_content = f.read()
            
            # S3に出力
            diff_detail_s3_key = f"{prefix}output/contract2-diff-result.txt"
            s3_client.put_object(
                Bucket=bucket_name,
                Key=diff_detail_s3_key,
                Body=diff_detail_content.encode('utf-8'),
                ContentType='text/plain; charset=utf-8'
            )
            
            print(f"Saved local file to: {local_file_path}")
            print(f"Saved diff details to: {diff_result_path}")
            print(f"Uploaded diff details to S3: s3://{bucket_name}/{diff_detail_s3_key}")
            
            output_s3_path = f"s3://{bucket_name}/{output_s3_key}"
            
            return {
                "message": "契約ファイルの差分計算が完了しました",
                "today_date": today_str,
                "yesterday_date": yesterday_str,
                "total_rows_today": len(today_df),
                "total_rows_yesterday": len(yesterday_df),
                "new_rows": len(only_in_today),
                "changed_rows": len(changed_indices),
                "diff_rows": len(diff_df),
                "output_file": output_s3_path,
                "debug_info": {
                    "detailed_diffs_found": len(detailed_diff_info),
                    "sample_differences": detailed_diff_info[:3] if detailed_diff_info else [],
                    "excluded_columns": "Column at index 20 excluded from comparison (original 22nd column)"
                }
            }
        else:
            raise HTTPException(status_code=400, detail="ファイルにカラムがありません")
            
    except (GoogleAPICallError, NotFound) as e:
        print(f"BigQuery Error: {e}")
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        print(f"Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post('/gcp/pallet-cloud/contract_tenant-diff')
def contract_tenant_diff(params: DaysAgoParams = None):
    try:
        # 除外するコントラクトIDのリストを読み込む
        exclude_contracts = []
        try:
            with open('./exclude_contracts.txt', 'r') as f:
                exclude_contracts = [line.strip() for line in f if line.strip()]
            print(f"Loaded {len(exclude_contracts)} contract IDs to exclude")
        except Exception as e:
            print(f"Warning: Could not load exclude_contracts.txt: {e}")
            
        # パラメータがない場合はデフォルト値を使用
        days_ago = 0 if params is None else params.days_ago
        
        # 現在の日付と指定された日数前の日付をyyyymmdd形式で取得
        today = datetime.datetime.now()
        base_date = today - datetime.timedelta(days=days_ago)
        yesterday = base_date - datetime.timedelta(days=1)
        today_str = base_date.strftime("%Y%m%d")
        yesterday_str = yesterday.strftime("%Y%m%d")
        
        # S3バケットとプレフィックス
        bucket_name = "adi-external-integration"
        prefix = "pallet-cloud/prod/"
        
        # 今日と昨日のファイルパス
        today_file = f"{today_str}_PC_contract_tenant.csv"
        yesterday_file = f"{yesterday_str}_PC_contract_tenant.csv"
        
        today_s3_path = f"{prefix}{today_file}"
        yesterday_s3_path = f"{prefix}{yesterday_file}"
        
        # S3クライアントの初期化
        s3_client = boto3.client('s3')
        
        # 一時ファイルパス
        temp_today_file = f"/tmp/{today_file}"
        temp_yesterday_file = f"/tmp/{yesterday_file}"
        
        # S3からファイルをダウンロード
        try:
            s3_client.download_file(bucket_name, today_s3_path, temp_today_file)
            print(f"Downloaded today's file from s3://{bucket_name}/{today_s3_path}")
        except Exception as e:
            print(f"Error downloading today's file: {e}")
            raise HTTPException(status_code=404, detail=f"Today's file not found: {today_file}")
            
        try:
            s3_client.download_file(bucket_name, yesterday_s3_path, temp_yesterday_file)
            print(f"Downloaded yesterday's file from s3://{bucket_name}/{yesterday_s3_path}")
        except Exception as e:
            print(f"Error downloading yesterday's file: {e}")
            raise HTTPException(status_code=404, detail=f"Yesterday's file not found: {yesterday_file}")
        
        # CSVファイルをDataFrameに読み込む（データクリーニング付き）
        def clean_dataframe(df):
            """DataFrameの文字列カラムをクリーニング"""
            for col in df.columns:
                if df[col].dtype == 'object':  # 文字列カラムの場合
                    # 空文字列をNaNに統一
                    df[col] = df[col].replace('', pd.NA)
                    # 先頭・末尾の空白を除去
                    df[col] = df[col].astype(str).str.strip()
                    # 'nan'文字列をNaNに変換
                    df[col] = df[col].replace('nan', pd.NA)
                    # 空白のみの文字列をNaNに変換
                    df[col] = df[col].replace(r'^\s*$', pd.NA, regex=True)
            return df
        
        today_df = pd.read_csv(temp_today_file, encoding='utf-8', keep_default_na=True, na_values=['', 'NULL', 'null', 'None'])
        yesterday_df = pd.read_csv(temp_yesterday_file, encoding='utf-8', keep_default_na=True, na_values=['', 'NULL', 'null', 'None'])
        
        # データクリーニング
        today_df = clean_dataframe(today_df)
        yesterday_df = clean_dataframe(yesterday_df)
        
        print(f"Today's file has {len(today_df)} rows")
        print(f"Yesterday's file has {len(yesterday_df)} rows")
        
        # 除外対象のcontract_idに一致する行を除外
        if len(today_df.columns) > 2 and exclude_contracts:
            contract_id_column = today_df.columns[2]  # 0-indexで2は3カラム目
            
            # デバッグ情報を出力
            print(f"Contract ID column: {contract_id_column}")
            print(f"Contract ID column dtype: {today_df[contract_id_column].dtype}")
            print(f"Sample contract IDs from today's file: {today_df[contract_id_column].head().tolist()}")
            print(f"Exclude contracts list (first 5): {exclude_contracts[:5]}")
            print(f"Exclude contracts types: {[type(x) for x in exclude_contracts[:5]]}")
            
            # 型変換: contract_idと除外リストを文字列に統一
            today_df[contract_id_column] = today_df[contract_id_column].astype(str)
            exclude_contracts_str = [str(x).strip() for x in exclude_contracts]
            
            # フィルタリング前の行数
            before_count = len(today_df)
            today_df = today_df[~today_df[contract_id_column].isin(exclude_contracts_str)]
            after_count = len(today_df)
            excluded_count = before_count - after_count
            
            print(f"Today's file: {before_count} -> {after_count} rows (excluded {excluded_count} rows)")
        
        if len(yesterday_df.columns) > 2 and exclude_contracts:
            contract_id_column = yesterday_df.columns[2]  # 0-indexで2は3カラム目
            
            # 型変換: contract_idと除外リストを文字列に統一
            yesterday_df[contract_id_column] = yesterday_df[contract_id_column].astype(str)
            exclude_contracts_str = [str(x).strip() for x in exclude_contracts]
            
            # フィルタリング前の行数
            before_count = len(yesterday_df)
            yesterday_df = yesterday_df[~yesterday_df[contract_id_column].isin(exclude_contracts_str)]
            after_count = len(yesterday_df)
            excluded_count = before_count - after_count
            
            print(f"Yesterday's file: {before_count} -> {after_count} rows (excluded {excluded_count} rows)")
        
        # 重複検知と除去機能を追加
        def detect_and_remove_duplicates(df, file_name):
            """重複を検知し、除去する"""
            if len(df.columns) == 0:
                return df, []
            
            primary_key = df.columns[0]
            duplicates = df[df.duplicated(subset=[primary_key], keep=False)]
            duplicate_keys = []
            
            if not duplicates.empty:
                duplicate_keys = duplicates[primary_key].unique().tolist()
                print(f"Found {len(duplicate_keys)} duplicate keys in {file_name}: {duplicate_keys}")
                # 重複を除去（最初の行を保持）
                df_deduped = df.drop_duplicates(subset=[primary_key], keep='first')
                print(f"After deduplication: {len(df_deduped)} rows (removed {len(df) - len(df_deduped)} duplicates)")
                return df_deduped, duplicate_keys
            else:
                print(f"No duplicates found in {file_name}")
                return df, duplicate_keys
        
        # 重複検知と除去を実行
        today_df, today_duplicates = detect_and_remove_duplicates(today_df, "today's file")
        yesterday_df, yesterday_duplicates = detect_and_remove_duplicates(yesterday_df, "yesterday's file")
        
        # 元のカラム順序を保存
        original_columns = today_df.columns.tolist()
        
        # 主キーを特定（1カラム目を主キーとする）
        if len(today_df.columns) > 1:
            primary_key = today_df.columns[0]
            
            # **最適化1: インデックスを設定して高速化**
            today_df_indexed = today_df.set_index(primary_key)
            yesterday_df_indexed = yesterday_df.set_index(primary_key)
            
            # **最適化2: 集合演算を使用**
            today_keys = set(today_df_indexed.index)
            yesterday_keys = set(yesterday_df_indexed.index)
            
            # 1. 今日のファイルにしかない行を抽出
            only_today_keys = today_keys - yesterday_keys
            only_in_today = today_df_indexed.loc[list(only_today_keys)] if only_today_keys else pd.DataFrame()
            
            # 2. 昨日のファイルにしかない行を抽出（削除された行）
            only_yesterday_keys = yesterday_keys - today_keys
            only_in_yesterday = yesterday_df_indexed.loc[list(only_yesterday_keys)] if only_yesterday_keys else pd.DataFrame()
            
            # 削除された行に2列目に1を挿入
            if not only_in_yesterday.empty and len(only_in_yesterday.columns) > 0:
                # 2列目の名前を取得（indexの次の列）
                second_column = only_in_yesterday.columns[0]
                # 2列目に1を設定
                only_in_yesterday[second_column] = 1
            
            # 3. 共通キーを取得
            common_keys = today_keys & yesterday_keys
            print(f"Processing {len(common_keys)} common keys for differences")
            
            # **最適化3: ベクトル化された比較**
            diff_details = []
            changed_indices = set()  # setを使用して重複排除を効率化
            detailed_diff_info = []  # デバッグ用の詳細情報
            
            if common_keys:
                # 共通キーのデータを一括で取得
                today_common = today_df_indexed.loc[list(common_keys)]
                yesterday_common = yesterday_df_indexed.loc[list(common_keys)]
                
                # **最適化4: カラムごとに一括比較（改良版）**
                for col in today_common.columns:
                    if col in yesterday_common.columns:
                        today_vals = today_common[col]
                        yesterday_vals = yesterday_common[col]
                        
                        # 数値型の場合は差分が1以上あるかチェック
                        if pd.api.types.is_numeric_dtype(today_vals):
                            # 両方がNaNでない場合の数値差分チェック
                            both_not_nan = pd.notna(today_vals) & pd.notna(yesterday_vals)
                            diff_mask = both_not_nan & (abs(today_vals - yesterday_vals) >= 1)
                            # 片方だけがNaNの場合もチェック
                            nan_diff_mask = today_vals.isna() != yesterday_vals.isna()
                            combined_mask = diff_mask | nan_diff_mask
                        else:
                            # 文字列型の場合：改良された比較ロジック
                            # NaNの扱いを統一（両方がNaNの場合は差分なし）
                            both_nan = pd.isna(today_vals) & pd.isna(yesterday_vals)
                            both_not_nan = pd.notna(today_vals) & pd.notna(yesterday_vals)
                            one_nan = pd.isna(today_vals) != pd.isna(yesterday_vals)
                            
                            # 両方が値を持つ場合の比較
                            value_diff = both_not_nan & (today_vals.astype(str) != yesterday_vals.astype(str))
                            
                            # 差分があるのは：値が異なる場合 OR 片方だけがNaNの場合
                            combined_mask = value_diff | one_nan
                        
                        # 差分があるキーをsetに追加（自動重複排除）
                        different_keys = combined_mask[combined_mask].index.tolist()
                        if different_keys:
                            changed_indices.update(different_keys)
                            
                            # デバッグ用：差分がある全件の詳細情報を記録
                            for key in different_keys:
                                today_val = today_vals.loc[key] if key in today_vals.index else 'KEY_NOT_FOUND'
                                yesterday_val = yesterday_vals.loc[key] if key in yesterday_vals.index else 'KEY_NOT_FOUND'
                                detailed_diff_info.append({
                                    'key': key,
                                    'column': col,
                                    'today_value': repr(today_val),
                                    'yesterday_value': repr(yesterday_val),
                                    'today_type': type(today_val).__name__,
                                    'yesterday_type': type(yesterday_val).__name__
                                })
                
                # setをリストに変換
                changed_indices = list(changed_indices)
                
                # 差分詳細記録
                for key in list(changed_indices):
                    diff_details.append({"key": key, "differences": ["Changes detected"]})
            
            # 変更された行を取得
            changed_rows = today_df_indexed.loc[changed_indices] if changed_indices else pd.DataFrame()
            
            # 差分ファイルを作成（今日にしかない行 + 変更された行 + 削除された行）
            diff_frames = []
            if not only_in_today.empty:
                # インデックスをリセットして元のカラム順序を復元
                only_in_today_reset = only_in_today.reset_index()
                only_in_today_ordered = only_in_today_reset[original_columns]
                diff_frames.append(only_in_today_ordered)
            
            if not changed_rows.empty:
                # インデックスをリセットして元のカラム順序を復元
                changed_rows_reset = changed_rows.reset_index()
                changed_rows_ordered = changed_rows_reset[original_columns]
                diff_frames.append(changed_rows_ordered)
            
            if not only_in_yesterday.empty:
                # インデックスをリセットして元のカラム順序を復元
                only_in_yesterday_reset = only_in_yesterday.reset_index()
                only_in_yesterday_ordered = only_in_yesterday_reset[original_columns]
                diff_frames.append(only_in_yesterday_ordered)
            
            if diff_frames:
                diff_df = pd.concat(diff_frames, ignore_index=True)
            else:
                diff_df = pd.DataFrame()
            
            # 差分ファイルをCSVに変換
            csv_buffer = io.StringIO()
            diff_df.to_csv(csv_buffer, index=False)
            
            # S3に出力（UTF-8エンコーディングを明示的に指定）
            output_s3_key = f"{prefix}output/contract_tenant.csv"
            s3_client.put_object(
                Bucket=bucket_name,
                Key=output_s3_key,
                Body=csv_buffer.getvalue().encode('utf-8'),
                ContentType='text/csv; charset=utf-8'
            )
            
            # ローカルにも保存
            local_dir = f"./pallet_cloud/{today_str}"
            os.makedirs(local_dir, exist_ok=True)
            local_file_path = f"{local_dir}/contract_tenant.csv"
            
            # CSVファイルを保存
            with open(local_file_path, 'w', encoding='utf-8') as f:
                f.write(csv_buffer.getvalue())
            
            # 差分詳細をテキストファイルに出力（デバッグ情報付き）
            diff_result_path = f"{local_dir}/contract_tenant-diff-result.txt"
            with open(diff_result_path, 'w', encoding='utf-8') as f:
                f.write(f"Contract Tenant Diff Results - {today_str} vs {yesterday_str}\n")
                f.write(f"Total rows in today's file: {len(today_df)}\n")
                f.write(f"Total rows in yesterday's file: {len(yesterday_df)}\n")
                
                # 重複情報を追加
                f.write(f"Duplicate keys in today's file: {len(today_duplicates)} - {today_duplicates}\n")
                f.write(f"Duplicate keys in yesterday's file: {len(yesterday_duplicates)} - {yesterday_duplicates}\n")
                
                f.write(f"New rows (only in today's file): {len(only_in_today)}\n")
                f.write(f"Deleted rows (only in yesterday's file): {len(only_in_yesterday)}\n")
                f.write(f"Changed rows: {len(changed_rows)}\n\n")
                
                # 重複詳細を出力
                if today_duplicates or yesterday_duplicates:
                    f.write("=== DUPLICATE INFORMATION ===\n")
                    if today_duplicates:
                        f.write(f"Today's file had {len(today_duplicates)} duplicate keys: {today_duplicates}\n")
                    if yesterday_duplicates:
                        f.write(f"Yesterday's file had {len(yesterday_duplicates)} duplicate keys: {yesterday_duplicates}\n")
                    f.write("Note: Duplicates were automatically removed (keeping first occurrence)\n\n")
                
                if len(only_in_today) > 0:
                    f.write("=== NEW ROWS ===\n")
                    for key in only_in_today.index:
                        f.write(f"New row with key: {key}\n")
                    f.write("\n")
                
                if len(only_in_yesterday) > 0:
                    f.write("=== DELETED ROWS ===\n")
                    for key in only_in_yesterday.index:
                        f.write(f"Deleted row with key: {key}\n")
                    f.write("\n")
                
                if len(diff_details) > 0:
                    f.write("=== CHANGED ROWS ===\n")
                    for row_diff in diff_details:
                        f.write(f"Row with key '{row_diff['key']}' has differences\n")
                    f.write("\n")
                    
                    # デバッグ用の詳細情報を出力
                    if detailed_diff_info:
                        f.write(f"=== DETAILED DIFF INFO (All {len(detailed_diff_info)} differences) ===\n")
                        for info in detailed_diff_info:
                            f.write(f"Key: {info['key']}, Column: {info['column']}\n")
                            f.write(f"  Today: {info['today_value']} (type: {info['today_type']})\n")
                            f.write(f"  Yesterday: {info['yesterday_value']} (type: {info['yesterday_type']})\n")
                            f.write("\n")
                else:
                    f.write("No differences found in existing rows.\n")
            
            # 差分詳細ファイルをS3にもアップロード
            with open(diff_result_path, 'r', encoding='utf-8') as f:
                diff_detail_content = f.read()
            
            # S3に出力
            diff_detail_s3_key = f"{prefix}output/contract_tenant-diff-result.txt"
            s3_client.put_object(
                Bucket=bucket_name,
                Key=diff_detail_s3_key,
                Body=diff_detail_content.encode('utf-8'),
                ContentType='text/plain; charset=utf-8'
            )
            
            print(f"Saved local file to: {local_file_path}")
            print(f"Saved diff details to: {diff_result_path}")
            print(f"Uploaded diff details to S3: s3://{bucket_name}/{diff_detail_s3_key}")
            
            output_s3_path = f"s3://{bucket_name}/{output_s3_key}"
            
            return {
                "message": "契約テナントファイルの差分計算が完了しました",
                "today_date": today_str,
                "yesterday_date": yesterday_str,
                "total_rows_today": len(today_df),
                "total_rows_yesterday": len(yesterday_df),
                "today_duplicates": len(today_duplicates),
                "yesterday_duplicates": len(yesterday_duplicates),
                "duplicate_keys_today": today_duplicates,
                "duplicate_keys_yesterday": yesterday_duplicates,
                "new_rows": len(only_in_today),
                "deleted_rows": len(only_in_yesterday),
                "changed_rows": len(changed_indices),
                "diff_rows": len(diff_df),
                "output_file": output_s3_path,
                "debug_info": {
                    "detailed_diffs_found": len(detailed_diff_info),
                    "sample_differences": detailed_diff_info[:3] if detailed_diff_info else [],
                    "note": "All columns included in comparison (no excluded columns)"
                }
            }
        else:
            raise HTTPException(status_code=400, detail="ファイルにカラムがありません")
            
    except (GoogleAPICallError, NotFound) as e:
        print(f"BigQuery Error: {e}")
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        print(f"Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post('/gcp/pallet-cloud/contract_resident-diff')
def contract_resident_diff(params: DaysAgoParams = None):
    try:
        # 除外するコントラクトIDのリストを読み込む
        exclude_contracts = []
        try:
            with open('./exclude_contracts.txt', 'r') as f:
                exclude_contracts = [line.strip() for line in f if line.strip()]
            print(f"Loaded {len(exclude_contracts)} contract IDs to exclude")
        except Exception as e:
            print(f"Warning: Could not load exclude_contracts.txt: {e}")
            
        # パラメータがない場合はデフォルト値を使用
        days_ago = 0 if params is None else params.days_ago
        
        # 現在の日付と指定された日数前の日付をyyyymmdd形式で取得
        today = datetime.datetime.now()
        base_date = today - datetime.timedelta(days=days_ago)
        yesterday = base_date - datetime.timedelta(days=1)
        today_str = base_date.strftime("%Y%m%d")
        yesterday_str = yesterday.strftime("%Y%m%d")
        
        # S3バケットとプレフィックス
        bucket_name = "adi-external-integration"
        prefix = "pallet-cloud/prod/"
        
        # 今日と昨日のファイルパス
        today_file = f"{today_str}_PC_contract_resident.csv"
        yesterday_file = f"{yesterday_str}_PC_contract_resident.csv"
        
        today_s3_path = f"{prefix}{today_file}"
        yesterday_s3_path = f"{prefix}{yesterday_file}"
        
        # S3クライアントの初期化
        s3_client = boto3.client('s3')
        
        # 一時ファイルパス
        temp_today_file = f"/tmp/{today_file}"
        temp_yesterday_file = f"/tmp/{yesterday_file}"
        
        # S3からファイルをダウンロード
        try:
            s3_client.download_file(bucket_name, today_s3_path, temp_today_file)
            print(f"Downloaded today's file from s3://{bucket_name}/{today_s3_path}")
        except Exception as e:
            print(f"Error downloading today's file: {e}")
            raise HTTPException(status_code=404, detail=f"Today's file not found: {today_file}")
            
        try:
            s3_client.download_file(bucket_name, yesterday_s3_path, temp_yesterday_file)
            print(f"Downloaded yesterday's file from s3://{bucket_name}/{yesterday_s3_path}")
        except Exception as e:
            print(f"Error downloading yesterday's file: {e}")
            raise HTTPException(status_code=404, detail=f"Yesterday's file not found: {yesterday_file}")
        
        # CSVファイルをDataFrameに読み込む（データクリーニング付き）
        def clean_dataframe(df):
            """DataFrameの文字列カラムをクリーニング"""
            for col in df.columns:
                if df[col].dtype == 'object':  # 文字列カラムの場合
                    # 空文字列をNaNに統一
                    df[col] = df[col].replace('', pd.NA)
                    # 先頭・末尾の空白を除去
                    df[col] = df[col].astype(str).str.strip()
                    # 'nan'文字列をNaNに変換
                    df[col] = df[col].replace('nan', pd.NA)
                    # 空白のみの文字列をNaNに変換
                    df[col] = df[col].replace(r'^\s*$', pd.NA, regex=True)
            return df
        
        today_df = pd.read_csv(temp_today_file, encoding='utf-8', keep_default_na=True, na_values=['', 'NULL', 'null', 'None'])
        yesterday_df = pd.read_csv(temp_yesterday_file, encoding='utf-8', keep_default_na=True, na_values=['', 'NULL', 'null', 'None'])
        
        # データクリーニング
        today_df = clean_dataframe(today_df)
        yesterday_df = clean_dataframe(yesterday_df)
        
        print(f"Today's file has {len(today_df)} rows")
        print(f"Yesterday's file has {len(yesterday_df)} rows")
        
        # 除外対象のcontract_idに一致する行を除外
        if len(today_df.columns) > 2 and exclude_contracts:
            contract_id_column = today_df.columns[2]  # 0-indexで2は3カラム目
            
            # デバッグ情報を出力
            print(f"Contract ID column: {contract_id_column}")
            print(f"Contract ID column dtype: {today_df[contract_id_column].dtype}")
            print(f"Sample contract IDs from today's file: {today_df[contract_id_column].head().tolist()}")
            print(f"Exclude contracts list (first 5): {exclude_contracts[:5]}")
            print(f"Exclude contracts types: {[type(x) for x in exclude_contracts[:5]]}")
            
            # 型変換: contract_idと除外リストを文字列に統一
            today_df[contract_id_column] = today_df[contract_id_column].astype(str)
            exclude_contracts_str = [str(x).strip() for x in exclude_contracts]
            
            # フィルタリング前の行数
            before_count = len(today_df)
            today_df = today_df[~today_df[contract_id_column].isin(exclude_contracts_str)]
            after_count = len(today_df)
            excluded_count = before_count - after_count
            
            print(f"Today's file: {before_count} -> {after_count} rows (excluded {excluded_count} rows)")
        
        if len(yesterday_df.columns) > 2 and exclude_contracts:
            contract_id_column = yesterday_df.columns[2]  # 0-indexで2は3カラム目
            
            # 型変換: contract_idと除外リストを文字列に統一
            yesterday_df[contract_id_column] = yesterday_df[contract_id_column].astype(str)
            exclude_contracts_str = [str(x).strip() for x in exclude_contracts]
            
            # フィルタリング前の行数
            before_count = len(yesterday_df)
            yesterday_df = yesterday_df[~yesterday_df[contract_id_column].isin(exclude_contracts_str)]
            after_count = len(yesterday_df)
            excluded_count = before_count - after_count
            
            print(f"Yesterday's file: {before_count} -> {after_count} rows (excluded {excluded_count} rows)")
        
        # 重複検知と除去機能を追加
        def detect_and_remove_duplicates(df, file_name):
            """重複を検知し、除去する"""
            if len(df.columns) == 0:
                return df, []
            
            primary_key = df.columns[0]
            duplicates = df[df.duplicated(subset=[primary_key], keep=False)]
            duplicate_keys = []
            
            if not duplicates.empty:
                duplicate_keys = duplicates[primary_key].unique().tolist()
                print(f"Found {len(duplicate_keys)} duplicate keys in {file_name}: {duplicate_keys}")
                # 重複を除去（最初の行を保持）
                df_deduped = df.drop_duplicates(subset=[primary_key], keep='first')
                print(f"After deduplication: {len(df_deduped)} rows (removed {len(df) - len(df_deduped)} duplicates)")
                return df_deduped, duplicate_keys
            else:
                print(f"No duplicates found in {file_name}")
                return df, duplicate_keys
        
        # 重複検知と除去を実行
        today_df, today_duplicates = detect_and_remove_duplicates(today_df, "today's file")
        yesterday_df, yesterday_duplicates = detect_and_remove_duplicates(yesterday_df, "yesterday's file")
        
        # 元のカラム順序を保存
        original_columns = today_df.columns.tolist()
        
        # 主キーを特定（1カラム目を主キーとする）
        if len(today_df.columns) > 1:
            primary_key = today_df.columns[0]
            
            # **最適化1: インデックスを設定して高速化**
            today_df_indexed = today_df.set_index(primary_key)
            yesterday_df_indexed = yesterday_df.set_index(primary_key)
            
            # **最適化2: 集合演算を使用**
            today_keys = set(today_df_indexed.index)
            yesterday_keys = set(yesterday_df_indexed.index)
            
            # 1. 今日のファイルにしかない行を抽出
            only_today_keys = today_keys - yesterday_keys
            only_in_today = today_df_indexed.loc[list(only_today_keys)] if only_today_keys else pd.DataFrame()
            
            # 2. 昨日のファイルにしかない行を抽出（削除された行）
            only_yesterday_keys = yesterday_keys - today_keys
            only_in_yesterday = yesterday_df_indexed.loc[list(only_yesterday_keys)] if only_yesterday_keys else pd.DataFrame()
            
            # 削除された行に2列目に1を挿入
            if not only_in_yesterday.empty and len(only_in_yesterday.columns) > 0:
                # 2列目の名前を取得（indexの次の列）
                second_column = only_in_yesterday.columns[0]
                # 2列目に1を設定
                only_in_yesterday[second_column] = 1
            
            # 3. 共通キーを取得
            common_keys = today_keys & yesterday_keys
            print(f"Processing {len(common_keys)} common keys for differences")
            
            # **最適化3: ベクトル化された比較**
            diff_details = []
            changed_indices = set()  # setを使用して重複排除を効率化
            detailed_diff_info = []  # デバッグ用の詳細情報
            
            if common_keys:
                # 共通キーのデータを一括で取得
                today_common = today_df_indexed.loc[list(common_keys)]
                yesterday_common = yesterday_df_indexed.loc[list(common_keys)]
                
                # **最適化4: カラムごとに一括比較（改良版）**
                for col in today_common.columns:
                    if col in yesterday_common.columns:
                        today_vals = today_common[col]
                        yesterday_vals = yesterday_common[col]
                        
                        # 数値型の場合は差分が1以上あるかチェック
                        if pd.api.types.is_numeric_dtype(today_vals):
                            # 両方がNaNでない場合の数値差分チェック
                            both_not_nan = pd.notna(today_vals) & pd.notna(yesterday_vals)
                            diff_mask = both_not_nan & (abs(today_vals - yesterday_vals) >= 1)
                            # 片方だけがNaNの場合もチェック
                            nan_diff_mask = today_vals.isna() != yesterday_vals.isna()
                            combined_mask = diff_mask | nan_diff_mask
                        else:
                            # 文字列型の場合：改良された比較ロジック
                            # NaNの扱いを統一（両方がNaNの場合は差分なし）
                            both_nan = pd.isna(today_vals) & pd.isna(yesterday_vals)
                            both_not_nan = pd.notna(today_vals) & pd.notna(yesterday_vals)
                            one_nan = pd.isna(today_vals) != pd.isna(yesterday_vals)
                            
                            # 両方が値を持つ場合の比較
                            value_diff = both_not_nan & (today_vals.astype(str) != yesterday_vals.astype(str))
                            
                            # 差分があるのは：値が異なる場合 OR 片方だけがNaNの場合
                            combined_mask = value_diff | one_nan
                        
                        # 差分があるキーをsetに追加（自動重複排除）
                        different_keys = combined_mask[combined_mask].index.tolist()
                        if different_keys:
                            changed_indices.update(different_keys)
                            
                            # デバッグ用：差分がある全件の詳細情報を記録
                            for key in different_keys:
                                today_val = today_vals.loc[key] if key in today_vals.index else 'KEY_NOT_FOUND'
                                yesterday_val = yesterday_vals.loc[key] if key in yesterday_vals.index else 'KEY_NOT_FOUND'
                                detailed_diff_info.append({
                                    'key': key,
                                    'column': col,
                                    'today_value': repr(today_val),
                                    'yesterday_value': repr(yesterday_val),
                                    'today_type': type(today_val).__name__,
                                    'yesterday_type': type(yesterday_val).__name__
                                })
                
                # setをリストに変換
                changed_indices = list(changed_indices)
                
                # 差分詳細記録
                for key in list(changed_indices):
                    diff_details.append({"key": key, "differences": ["Changes detected"]})
            
            # 変更された行を取得
            changed_rows = today_df_indexed.loc[changed_indices] if changed_indices else pd.DataFrame()
            
            # 差分ファイルを作成（今日にしかない行 + 変更された行 + 削除された行）
            diff_frames = []
            if not only_in_today.empty:
                # インデックスをリセットして元のカラム順序を復元
                only_in_today_reset = only_in_today.reset_index()
                only_in_today_ordered = only_in_today_reset[original_columns]
                diff_frames.append(only_in_today_ordered)
            
            if not changed_rows.empty:
                # インデックスをリセットして元のカラム順序を復元
                changed_rows_reset = changed_rows.reset_index()
                changed_rows_ordered = changed_rows_reset[original_columns]
                diff_frames.append(changed_rows_ordered)
            
            if not only_in_yesterday.empty:
                # インデックスをリセットして元のカラム順序を復元
                only_in_yesterday_reset = only_in_yesterday.reset_index()
                only_in_yesterday_ordered = only_in_yesterday_reset[original_columns]
                diff_frames.append(only_in_yesterday_ordered)
            
            if diff_frames:
                diff_df = pd.concat(diff_frames, ignore_index=True)
            else:
                diff_df = pd.DataFrame()
            
            # 差分ファイルをCSVに変換
            csv_buffer = io.StringIO()
            diff_df.to_csv(csv_buffer, index=False)
            
            # S3に出力（UTF-8エンコーディングを明示的に指定）
            output_s3_key = f"{prefix}output/contract_resident.csv"
            s3_client.put_object(
                Bucket=bucket_name,
                Key=output_s3_key,
                Body=csv_buffer.getvalue().encode('utf-8'),
                ContentType='text/csv; charset=utf-8'
            )
            
            # ローカルにも保存
            local_dir = f"./pallet_cloud/{today_str}"
            os.makedirs(local_dir, exist_ok=True)
            local_file_path = f"{local_dir}/contract_resident.csv"
            
            # CSVファイルを保存
            with open(local_file_path, 'w', encoding='utf-8') as f:
                f.write(csv_buffer.getvalue())
            
            # 差分詳細をテキストファイルに出力（デバッグ情報付き）
            diff_result_path = f"{local_dir}/contract_resident-diff-result.txt"
            with open(diff_result_path, 'w', encoding='utf-8') as f:
                f.write(f"Contract Resident Diff Results - {today_str} vs {yesterday_str}\n")
                f.write(f"Total rows in today's file: {len(today_df)}\n")
                f.write(f"Total rows in yesterday's file: {len(yesterday_df)}\n")
                
                # 重複情報を追加
                f.write(f"Duplicate keys in today's file: {len(today_duplicates)} - {today_duplicates}\n")
                f.write(f"Duplicate keys in yesterday's file: {len(yesterday_duplicates)} - {yesterday_duplicates}\n")
                
                f.write(f"New rows (only in today's file): {len(only_in_today)}\n")
                f.write(f"Deleted rows (only in yesterday's file): {len(only_in_yesterday)}\n")
                f.write(f"Changed rows: {len(changed_rows)}\n\n")
                
                # 重複詳細を出力
                if today_duplicates or yesterday_duplicates:
                    f.write("=== DUPLICATE INFORMATION ===\n")
                    if today_duplicates:
                        f.write(f"Today's file had {len(today_duplicates)} duplicate keys: {today_duplicates}\n")
                    if yesterday_duplicates:
                        f.write(f"Yesterday's file had {len(yesterday_duplicates)} duplicate keys: {yesterday_duplicates}\n")
                    f.write("Note: Duplicates were automatically removed (keeping first occurrence)\n\n")
                
                if len(only_in_today) > 0:
                    f.write("=== NEW ROWS ===\n")
                    for key in only_in_today.index:
                        f.write(f"New row with key: {key}\n")
                    f.write("\n")
                
                if len(only_in_yesterday) > 0:
                    f.write("=== DELETED ROWS ===\n")
                    for key in only_in_yesterday.index:
                        f.write(f"Deleted row with key: {key}\n")
                    f.write("\n")
                
                if len(diff_details) > 0:
                    f.write("=== CHANGED ROWS ===\n")
                    for row_diff in diff_details:
                        f.write(f"Row with key '{row_diff['key']}' has differences\n")
                    f.write("\n")
                    
                    # デバッグ用の詳細情報を出力
                    if detailed_diff_info:
                        f.write(f"=== DETAILED DIFF INFO (All {len(detailed_diff_info)} differences) ===\n")
                        for info in detailed_diff_info:
                            f.write(f"Key: {info['key']}, Column: {info['column']}\n")
                            f.write(f"  Today: {info['today_value']} (type: {info['today_type']})\n")
                            f.write(f"  Yesterday: {info['yesterday_value']} (type: {info['yesterday_type']})\n")
                            f.write("\n")
                else:
                    f.write("No differences found in existing rows.\n")
            
            # 差分詳細ファイルをS3にもアップロード
            with open(diff_result_path, 'r', encoding='utf-8') as f:
                diff_detail_content = f.read()
            
            # S3に出力
            diff_detail_s3_key = f"{prefix}output/contract_resident-diff-result.txt"
            s3_client.put_object(
                Bucket=bucket_name,
                Key=diff_detail_s3_key,
                Body=diff_detail_content.encode('utf-8'),
                ContentType='text/plain; charset=utf-8'
            )
            
            print(f"Saved local file to: {local_file_path}")
            print(f"Saved diff details to: {diff_result_path}")
            print(f"Uploaded diff details to S3: s3://{bucket_name}/{diff_detail_s3_key}")
            
            output_s3_path = f"s3://{bucket_name}/{output_s3_key}"
            
            return {
                "message": "契約居住者ファイルの差分計算が完了しました",
                "today_date": today_str,
                "yesterday_date": yesterday_str,
                "total_rows_today": len(today_df),
                "total_rows_yesterday": len(yesterday_df),
                "today_duplicates": len(today_duplicates),
                "yesterday_duplicates": len(yesterday_duplicates),
                "duplicate_keys_today": today_duplicates,
                "duplicate_keys_yesterday": yesterday_duplicates,
                "new_rows": len(only_in_today),
                "deleted_rows": len(only_in_yesterday),
                "changed_rows": len(changed_indices),
                "diff_rows": len(diff_df),
                "output_file": output_s3_path,
                "debug_info": {
                    "detailed_diffs_found": len(detailed_diff_info),
                    "sample_differences": detailed_diff_info[:3] if detailed_diff_info else [],
                    "note": "All columns included in comparison (no excluded columns)"
                }
            }
        else:
            raise HTTPException(status_code=400, detail="ファイルにカラムがありません")
            
    except (GoogleAPICallError, NotFound) as e:
        print(f"BigQuery Error: {e}")
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        print(f"Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post('/gcp/pallet-cloud/tenants-diff')
def tenants_diff(params: DaysAgoParams = None):
    try:
        # パラメータがない場合はデフォルト値を使用
        days_ago = 0 if params is None else params.days_ago
        
        # 現在の日付と指定された日数前の日付をyyyymmdd形式で取得
        today = datetime.datetime.now()
        base_date = today - datetime.timedelta(days=days_ago)
        yesterday = base_date - datetime.timedelta(days=1)
        today_str = base_date.strftime("%Y%m%d")
        yesterday_str = yesterday.strftime("%Y%m%d")
        
        # S3バケットとプレフィックス
        bucket_name = "adi-external-integration"
        prefix = "pallet-cloud/prod/"
        
        # 今日と昨日のファイルパス
        today_file = f"{today_str}_PC_tenants.csv"
        yesterday_file = f"{yesterday_str}_PC_tenants.csv"
        
        today_s3_path = f"{prefix}{today_file}"
        yesterday_s3_path = f"{prefix}{yesterday_file}"
        
        # S3クライアントの初期化
        s3_client = boto3.client('s3')
        
        # 一時ファイルパス
        temp_today_file = f"/tmp/{today_file}"
        temp_yesterday_file = f"/tmp/{yesterday_file}"
        
        # S3からファイルをダウンロード
        try:
            s3_client.download_file(bucket_name, today_s3_path, temp_today_file)
            print(f"Downloaded today's file from s3://{bucket_name}/{today_s3_path}")
        except Exception as e:
            print(f"Error downloading today's file: {e}")
            raise HTTPException(status_code=404, detail=f"Today's file not found: {today_file}")
            
        try:
            s3_client.download_file(bucket_name, yesterday_s3_path, temp_yesterday_file)
            print(f"Downloaded yesterday's file from s3://{bucket_name}/{yesterday_s3_path}")
        except Exception as e:
            print(f"Error downloading yesterday's file: {e}")
            raise HTTPException(status_code=404, detail=f"Yesterday's file not found: {yesterday_file}")
        
        # CSVファイルをDataFrameに読み込む（データクリーニング付き）
        def clean_dataframe(df):
            """DataFrameの文字列カラムをクリーニング"""
            for col in df.columns:
                if df[col].dtype == 'object':  # 文字列カラムの場合
                    # 空文字列をNaNに統一
                    df[col] = df[col].replace('', pd.NA)
                    # 先頭・末尾の空白を除去
                    df[col] = df[col].astype(str).str.strip()
                    # 'nan'文字列をNaNに変換
                    df[col] = df[col].replace('nan', pd.NA)
                    # 空白のみの文字列をNaNに変換
                    df[col] = df[col].replace(r'^\s*$', pd.NA, regex=True)
            return df
        
        today_df = pd.read_csv(temp_today_file, encoding='utf-8', keep_default_na=True, na_values=['', 'NULL', 'null', 'None'])
        yesterday_df = pd.read_csv(temp_yesterday_file, encoding='utf-8', keep_default_na=True, na_values=['', 'NULL', 'null', 'None'])
        
        # データクリーニング
        today_df = clean_dataframe(today_df)
        yesterday_df = clean_dataframe(yesterday_df)
        
        print(f"Today's file has {len(today_df)} rows")
        print(f"Yesterday's file has {len(yesterday_df)} rows")
        
        # 元のカラム順序を保存
        original_columns = today_df.columns.tolist()
        
        # 主キーを特定（1カラム目を主キーとする）
        if len(today_df.columns) > 0:
            primary_key = today_df.columns[0]
            
            # 重複チェックと処理
            today_duplicates = today_df[primary_key].duplicated()
            yesterday_duplicates = yesterday_df[primary_key].duplicated()
            
            if today_duplicates.any() or yesterday_duplicates.any():
                print(f"Warning: Found duplicates in primary key. Today: {today_duplicates.sum()}, Yesterday: {yesterday_duplicates.sum()}")
                # 重複を除去（最初の値を保持）
                today_df = today_df[~today_duplicates]
                yesterday_df = yesterday_df[~yesterday_duplicates]
                print(f"After removing duplicates - Today: {len(today_df)} rows, Yesterday: {len(yesterday_df)} rows")
            
            # **最適化1: インデックスを設定して高速化**
            today_df_indexed = today_df.set_index(primary_key)
            yesterday_df_indexed = yesterday_df.set_index(primary_key)
            
            # **最適化2: 集合演算を使用**
            today_keys = set(today_df_indexed.index)
            yesterday_keys = set(yesterday_df_indexed.index)
            
            # 1. 今日のファイルにしかない行を抽出
            only_today_keys = today_keys - yesterday_keys
            only_in_today = today_df_indexed.loc[list(only_today_keys)] if only_today_keys else pd.DataFrame()
            
            # 2. 共通キーを取得
            common_keys = today_keys & yesterday_keys
            print(f"Processing {len(common_keys)} common keys for differences")
            
            # **最適化3: ベクトル化された比較**
            diff_details = []
            changed_indices = set()  # setを使用して重複排除を効率化
            detailed_diff_info = []  # デバッグ用の詳細情報
            
            if common_keys:
                # 共通キーのデータを一括で取得
                today_common = today_df_indexed.loc[list(common_keys)]
                yesterday_common = yesterday_df_indexed.loc[list(common_keys)]
                
                # 比較から除外するカラム（インデックス）
                exclude_indices = [37, 38, 49] # index分ずれるからこれで39, 40, 41
                include_columns = [col for i, col in enumerate(today_common.columns) if i not in exclude_indices]
                
                # **最適化4: カラムごとに一括比較（改良版）**
                for col in include_columns:
                    if col in today_common.columns and col in yesterday_common.columns:
                        today_vals = today_common[col]
                        yesterday_vals = yesterday_common[col]
                        
                        # 数値型の場合は差分が1以上あるかチェック
                        if pd.api.types.is_numeric_dtype(today_vals):
                            # 両方がNaNでない場合の数値差分チェック
                            both_not_nan = pd.notna(today_vals) & pd.notna(yesterday_vals)
                            diff_mask = both_not_nan & (abs(today_vals - yesterday_vals) >= 1)
                            # 片方だけがNaNの場合もチェック
                            nan_diff_mask = today_vals.isna() != yesterday_vals.isna()
                            combined_mask = diff_mask | nan_diff_mask
                        else:
                            # 文字列型の場合：改良された比較ロジック
                            # NaNの扱いを統一（両方がNaNの場合は差分なし）
                            both_nan = pd.isna(today_vals) & pd.isna(yesterday_vals)
                            both_not_nan = pd.notna(today_vals) & pd.notna(yesterday_vals)
                            one_nan = pd.isna(today_vals) != pd.isna(yesterday_vals)
                            
                            # 両方が値を持つ場合の比較
                            value_diff = both_not_nan & (today_vals.astype(str) != yesterday_vals.astype(str))
                            
                            # 差分があるのは：値が異なる場合 OR 片方だけがNaNの場合
                            combined_mask = value_diff | one_nan
                        
                        # 差分があるキーをsetに追加（自動重複排除）
                        different_keys = combined_mask[combined_mask].index.tolist()
                        if different_keys:
                            changed_indices.update(different_keys)
                            
                            # デバッグ用：差分がある全件の詳細情報を記録
                            for key in different_keys:
                                today_val = today_vals.loc[key] if key in today_vals.index else 'KEY_NOT_FOUND'
                                yesterday_val = yesterday_vals.loc[key] if key in yesterday_vals.index else 'KEY_NOT_FOUND'
                                detailed_diff_info.append({
                                    'key': key,
                                    'column': col,
                                    'today_value': repr(today_val),
                                    'yesterday_value': repr(yesterday_val),
                                    'today_type': type(today_val).__name__,
                                    'yesterday_type': type(yesterday_val).__name__
                                })
                
                # setをリストに変換
                changed_indices = list(changed_indices)
                
                # 差分詳細記録（最初の100件のみ）
                for key in list(changed_indices)[:100]:
                    diff_details.append({"key": key, "differences": ["Changes detected"]})
            
            # 変更された行を取得
            changed_rows = today_df_indexed.loc[changed_indices] if changed_indices else pd.DataFrame()
            
            # 差分ファイルを作成（今日にしかない行 + 変更された行）
            diff_frames = []
            if not only_in_today.empty:
                # インデックスをリセットして元のカラム順序を復元
                only_in_today_reset = only_in_today.reset_index()
                only_in_today_ordered = only_in_today_reset[original_columns]
                diff_frames.append(only_in_today_ordered)
            
            if not changed_rows.empty:
                # インデックスをリセットして元のカラム順序を復元
                changed_rows_reset = changed_rows.reset_index()
                changed_rows_ordered = changed_rows_reset[original_columns]
                diff_frames.append(changed_rows_ordered)
            
            if diff_frames:
                diff_df = pd.concat(diff_frames, ignore_index=True)
            else:
                diff_df = pd.DataFrame()
            
            # 差分ファイルをCSVに変換
            csv_buffer = io.StringIO()
            diff_df.to_csv(csv_buffer, index=False)
            
            # S3に出力（UTF-8エンコーディングを明示的に指定）
            output_s3_key = f"{prefix}output/tenant.csv"
            s3_client.put_object(
                Bucket=bucket_name,
                Key=output_s3_key,
                Body=csv_buffer.getvalue().encode('utf-8'),
                ContentType='text/csv; charset=utf-8'
            )
            
            # ローカルにも保存
            local_dir = f"./pallet_cloud/{today_str}"
            os.makedirs(local_dir, exist_ok=True)
            local_file_path = f"{local_dir}/tenant.csv"
            
            # CSVファイルを保存
            with open(local_file_path, 'w', encoding='utf-8') as f:
                f.write(csv_buffer.getvalue())
            
            # 差分詳細をテキストファイルに出力（デバッグ情報付き）
            diff_result_path = f"{local_dir}/tenant-diff-result.txt"
            with open(diff_result_path, 'w', encoding='utf-8') as f:
                f.write(f"Tenant Diff Results - {today_str} vs {yesterday_str}\n")
                f.write(f"Total rows in today's file: {len(today_df)}\n")
                f.write(f"Total rows in yesterday's file: {len(yesterday_df)}\n")
                f.write(f"New rows (only in today's file): {len(only_in_today)}\n")
                f.write(f"Changed rows: {len(changed_rows)}\n\n")
                
                if len(only_in_today) > 0:
                    f.write("=== NEW ROWS ===\n")
                    for key in only_in_today.index:
                        f.write(f"New row with key: {key}\n")
                    f.write("\n")
                
                if len(diff_details) > 0:
                    f.write("=== CHANGED ROWS ===\n")
                    for row_diff in diff_details:
                        f.write(f"Row with key '{row_diff['key']}' has differences\n")
                    f.write("\n")
                    
                    # デバッグ用の詳細情報を出力
                    if detailed_diff_info:
                        f.write(f"=== DETAILED DIFF INFO (All {len(detailed_diff_info)} differences) ===\n")
                        for info in detailed_diff_info:
                            f.write(f"Key: {info['key']}, Column: {info['column']}\n")
                            f.write(f"  Today: {info['today_value']} (type: {info['today_type']})\n")
                            f.write(f"  Yesterday: {info['yesterday_value']} (type: {info['yesterday_type']})\n")
                            f.write("\n")
                else:
                    f.write("No differences found in existing rows.\n")
            
            # 差分詳細ファイルをS3にもアップロード
            with open(diff_result_path, 'r', encoding='utf-8') as f:
                diff_detail_content = f.read()
            
            # S3に出力
            diff_detail_s3_key = f"{prefix}output/tenant-diff-result.txt"
            s3_client.put_object(
                Bucket=bucket_name,
                Key=diff_detail_s3_key,
                Body=diff_detail_content.encode('utf-8'),
                ContentType='text/plain; charset=utf-8'
            )
            
            print(f"Saved local file to: {local_file_path}")
            print(f"Saved diff details to: {diff_result_path}")
            print(f"Uploaded diff details to S3: s3://{bucket_name}/{diff_detail_s3_key}")
            
            output_s3_path = f"s3://{bucket_name}/{output_s3_key}"
            
            return {
                "message": "テナントファイルの差分計算が完了しました",
                "today_date": today_str,
                "yesterday_date": yesterday_str,
                "total_rows_today": len(today_df),
                "total_rows_yesterday": len(yesterday_df),
                "new_rows": len(only_in_today),
                "changed_rows": len(changed_indices),
                "diff_rows": len(diff_df),
                "output_file": output_s3_path,
                "debug_info": {
                    "detailed_diffs_found": len(detailed_diff_info),
                    "sample_differences": detailed_diff_info[:3] if detailed_diff_info else [],
                    "excluded_columns": "Columns at index 39, 40, 41 excluded from comparison"
                }
            }
        else:
            raise HTTPException(status_code=400, detail="ファイルにカラムがありません")
            
    except (GoogleAPICallError, NotFound) as e:
        print(f"BigQuery Error: {e}")
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        print(f"Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post('/gcp/pallet-cloud/building-diff')
def building_diff(params: DaysAgoParams = None):
    try:
        # パラメータがない場合はデフォルト値を使用
        days_ago = 0 if params is None else params.days_ago
        
        # 現在の日付と指定された日数前の日付をyyyymmdd形式で取得
        today = datetime.datetime.now()
        base_date = today - datetime.timedelta(days=days_ago)
        yesterday = base_date - datetime.timedelta(days=1)
        today_str = base_date.strftime("%Y%m%d")
        yesterday_str = yesterday.strftime("%Y%m%d")
        
        # S3バケットとプレフィックス
        bucket_name = "adi-external-integration"
        prefix = "pallet-cloud/prod/"
        
        # 今日と昨日のファイルパス
        today_file = f"{today_str}_PC_buildings.csv"
        yesterday_file = f"{yesterday_str}_PC_buildings.csv"
        
        today_s3_path = f"{prefix}{today_file}"
        yesterday_s3_path = f"{prefix}{yesterday_file}"
        
        # S3クライアントの初期化
        s3_client = boto3.client('s3')
        
        # 一時ファイルパス
        temp_today_file = f"/tmp/{today_file}"
        temp_yesterday_file = f"/tmp/{yesterday_file}"
        
        # S3からファイルをダウンロード
        try:
            s3_client.download_file(bucket_name, today_s3_path, temp_today_file)
            print(f"Downloaded today's file from s3://{bucket_name}/{today_s3_path}")
        except Exception as e:
            print(f"Error downloading today's file: {e}")
            raise HTTPException(status_code=404, detail=f"Today's file not found: {today_file}")
            
        try:
            s3_client.download_file(bucket_name, yesterday_s3_path, temp_yesterday_file)
            print(f"Downloaded yesterday's file from s3://{bucket_name}/{yesterday_s3_path}")
        except Exception as e:
            print(f"Error downloading yesterday's file: {e}")
            raise HTTPException(status_code=404, detail=f"Yesterday's file not found: {yesterday_file}")
        
        # CSVファイルをDataFrameに読み込む（データクリーニング付き）
        def clean_dataframe(df):
            """DataFrameの文字列カラムをクリーニング"""
            for col in df.columns:
                if df[col].dtype == 'object':  # 文字列カラムの場合
                    # 空文字列をNaNに統一
                    df[col] = df[col].replace('', pd.NA)
                    # 先頭・末尾の空白を除去
                    df[col] = df[col].astype(str).str.strip()
                    # 'nan'文字列をNaNに変換
                    df[col] = df[col].replace('nan', pd.NA)
                    # 空白のみの文字列をNaNに変換
                    df[col] = df[col].replace(r'^\s*$', pd.NA, regex=True)
            return df
        
        today_df = pd.read_csv(temp_today_file, encoding='utf-8', keep_default_na=True, na_values=['', 'NULL', 'null', 'None'])
        yesterday_df = pd.read_csv(temp_yesterday_file, encoding='utf-8', keep_default_na=True, na_values=['', 'NULL', 'null', 'None'])
        
        # データクリーニング
        today_df = clean_dataframe(today_df)
        yesterday_df = clean_dataframe(yesterday_df)
        
        print(f"Today's file has {len(today_df)} rows")
        print(f"Yesterday's file has {len(yesterday_df)} rows")
        
        # 元のカラム順序を保存
        original_columns = today_df.columns.tolist()
        
        # 主キーを特定（2カラム目を主キーとする）
        if len(today_df.columns) > 1:
            primary_key = today_df.columns[1]
            
            # **最適化1: インデックスを設定して高速化**
            today_df_indexed = today_df.set_index(primary_key)
            yesterday_df_indexed = yesterday_df.set_index(primary_key)
            
            # **最適化2: 集合演算を使用**
            today_keys = set(today_df_indexed.index)
            yesterday_keys = set(yesterday_df_indexed.index)
            
            # 1. 今日のファイルにしかない行を抽出
            only_today_keys = today_keys - yesterday_keys
            only_in_today = today_df_indexed.loc[list(only_today_keys)] if only_today_keys else pd.DataFrame()
            
            # 2. 共通キーを取得
            common_keys = today_keys & yesterday_keys
            print(f"Processing {len(common_keys)} common keys for differences")
            
            # **最適化3: ベクトル化された比較**
            diff_details = []
            changed_indices = set()  # setを使用して重複排除を効率化
            detailed_diff_info = []  # デバッグ用の詳細情報
            
            if common_keys:
                # 共通キーのデータを一括で取得
                today_common = today_df_indexed.loc[list(common_keys)]
                yesterday_common = yesterday_df_indexed.loc[list(common_keys)]
                
                # 比較から除外するカラム（インデックス）
                exclude_indices = [60, 61]
                include_columns = [col for i, col in enumerate(today_common.columns) if i not in exclude_indices]
                
                # **最適化4: カラムごとに一括比較（改良版）**
                for col in include_columns:
                    if col in today_common.columns and col in yesterday_common.columns:
                        today_vals = today_common[col]
                        yesterday_vals = yesterday_common[col]
                        
                        # 数値型の場合は差分が1以上あるかチェック
                        if pd.api.types.is_numeric_dtype(today_vals):
                            # 両方がNaNでない場合の数値差分チェック
                            both_not_nan = pd.notna(today_vals) & pd.notna(yesterday_vals)
                            diff_mask = both_not_nan & (abs(today_vals - yesterday_vals) >= 1)
                            # 片方だけがNaNの場合もチェック
                            nan_diff_mask = today_vals.isna() != yesterday_vals.isna()
                            combined_mask = diff_mask | nan_diff_mask
                        else:
                            # 文字列型の場合：改良された比較ロジック
                            # NaNの扱いを統一（両方がNaNの場合は差分なし）
                            both_nan = pd.isna(today_vals) & pd.isna(yesterday_vals)
                            both_not_nan = pd.notna(today_vals) & pd.notna(yesterday_vals)
                            one_nan = pd.isna(today_vals) != pd.isna(yesterday_vals)
                            
                            # 両方が値を持つ場合の比較
                            value_diff = both_not_nan & (today_vals.astype(str) != yesterday_vals.astype(str))
                            
                            # 差分があるのは：値が異なる場合 OR 片方だけがNaNの場合
                            combined_mask = value_diff | one_nan
                        
                        # 差分があるキーをsetに追加（自動重複排除）
                        different_keys = combined_mask[combined_mask].index.tolist()
                        if different_keys:
                            changed_indices.update(different_keys)
                            
                            # デバッグ用：差分がある全件の詳細情報を記録
                            for key in different_keys:
                                today_val = today_vals.loc[key] if key in today_vals.index else 'KEY_NOT_FOUND'
                                yesterday_val = yesterday_vals.loc[key] if key in yesterday_vals.index else 'KEY_NOT_FOUND'
                                detailed_diff_info.append({
                                    'key': key,
                                    'column': col,
                                    'today_value': repr(today_val),
                                    'yesterday_value': repr(yesterday_val),
                                    'today_type': type(today_val).__name__,
                                    'yesterday_type': type(yesterday_val).__name__
                                })
                
                # setをリストに変換
                changed_indices = list(changed_indices)
                
                # 差分詳細記録
                for key in list(changed_indices):
                    diff_details.append({"key": key, "differences": ["Changes detected"]})
            
            # 変更された行を取得
            changed_rows = today_df_indexed.loc[changed_indices] if changed_indices else pd.DataFrame()
            
            # 差分ファイルを作成（今日にしかない行 + 変更された行）
            diff_frames = []
            if not only_in_today.empty:
                # インデックスをリセットして元のカラム順序を復元
                only_in_today_reset = only_in_today.reset_index()
                only_in_today_ordered = only_in_today_reset[original_columns]
                diff_frames.append(only_in_today_ordered)
            
            if not changed_rows.empty:
                # インデックスをリセットして元のカラム順序を復元
                changed_rows_reset = changed_rows.reset_index()
                changed_rows_ordered = changed_rows_reset[original_columns]
                diff_frames.append(changed_rows_ordered)
            
            if diff_frames:
                diff_df = pd.concat(diff_frames, ignore_index=True)
            else:
                diff_df = pd.DataFrame()
            
            # 差分ファイルをCSVに変換
            csv_buffer = io.StringIO()
            diff_df.to_csv(csv_buffer, index=False)
            
            # S3に出力（UTF-8エンコーディングを明示的に指定）
            output_s3_key = f"{prefix}output/building.csv"
            s3_client.put_object(
                Bucket=bucket_name,
                Key=output_s3_key,
                Body=csv_buffer.getvalue().encode('utf-8'),
                ContentType='text/csv; charset=utf-8'
            )
            
            # ローカルにも保存
            local_dir = f"./pallet_cloud/{today_str}"
            os.makedirs(local_dir, exist_ok=True)
            local_file_path = f"{local_dir}/building.csv"
            
            # CSVファイルを保存
            with open(local_file_path, 'w', encoding='utf-8') as f:
                f.write(csv_buffer.getvalue())
            
            # 差分詳細をテキストファイルに出力（デバッグ情報付き）
            diff_result_path = f"{local_dir}/building-diff-result.txt"
            with open(diff_result_path, 'w', encoding='utf-8') as f:
                f.write(f"Building Diff Results - {today_str} vs {yesterday_str}\n")
                f.write(f"Total rows in today's file: {len(today_df)}\n")
                f.write(f"Total rows in yesterday's file: {len(yesterday_df)}\n")
                f.write(f"New rows (only in today's file): {len(only_in_today)}\n")
                f.write(f"Changed rows: {len(changed_rows)}\n\n")
                
                if len(only_in_today) > 0:
                    f.write("=== NEW ROWS ===\n")
                    for key in only_in_today.index:
                        f.write(f"New row with key: {key}\n")
                    f.write("\n")
                
                if len(diff_details) > 0:
                    f.write("=== CHANGED ROWS ===\n")
                    for row_diff in diff_details:
                        f.write(f"Row with key '{row_diff['key']}' has differences\n")
                    f.write("\n")
                    
                    # デバッグ用の詳細情報を出力
                    if detailed_diff_info:
                        f.write(f"=== DETAILED DIFF INFO (All {len(detailed_diff_info)} differences) ===\n")
                        for info in detailed_diff_info:
                            f.write(f"Key: {info['key']}, Column: {info['column']}\n")
                            f.write(f"  Today: {info['today_value']} (type: {info['today_type']})\n")
                            f.write(f"  Yesterday: {info['yesterday_value']} (type: {info['yesterday_type']})\n")
                            f.write("\n")
                else:
                    f.write("No differences found in existing rows.\n")
            
            # 差分詳細ファイルをS3にもアップロード
            with open(diff_result_path, 'r', encoding='utf-8') as f:
                diff_detail_content = f.read()
            
            # S3に出力
            diff_detail_s3_key = f"{prefix}output/building-diff-result.txt"
            s3_client.put_object(
                Bucket=bucket_name,
                Key=diff_detail_s3_key,
                Body=diff_detail_content.encode('utf-8'),
                ContentType='text/plain; charset=utf-8'
            )
            
            print(f"Saved local file to: {local_file_path}")
            print(f"Saved diff details to: {diff_result_path}")
            print(f"Uploaded diff details to S3: s3://{bucket_name}/{diff_detail_s3_key}")
            
            output_s3_path = f"s3://{bucket_name}/{output_s3_key}"
            
            return {
                "message": "建物ファイルの差分計算が完了しました",
                "today_date": today_str,
                "yesterday_date": yesterday_str,
                "total_rows_today": len(today_df),
                "total_rows_yesterday": len(yesterday_df),
                "new_rows": len(only_in_today),
                "changed_rows": len(changed_indices),
                "diff_rows": len(diff_df),
                "output_file": output_s3_path,
                "debug_info": {
                    "detailed_diffs_found": len(detailed_diff_info),
                    "sample_differences": detailed_diff_info[:3] if detailed_diff_info else [],
                    "excluded_columns": "Columns at index 60, 61 excluded from comparison"
                }
            }
        else:
            raise HTTPException(status_code=400, detail="ファイルにカラムがありません")
            
    except (GoogleAPICallError, NotFound) as e:
        print(f"BigQuery Error: {e}")
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        print(f"Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))
        

@router.post('/gcp/pallet-cloud/commitment-diff')
def commitment_diff(params: DaysAgoParams = None):
    try:
        # 除外するコントラクトIDのリストを読み込む
        exclude_contracts = []
        try:
            with open('./exclude_contracts.txt', 'r') as f:
                exclude_contracts = [line.strip() for line in f if line.strip()]
            print(f"Loaded {len(exclude_contracts)} contract IDs to exclude")
        except Exception as e:
            print(f"Warning: Could not load exclude_contracts.txt: {e}")
            
        # パラメータがない場合はデフォルト値を使用
        days_ago = 0 if params is None else params.days_ago
        
        # 現在の日付と指定された日数前の日付をyyyymmdd形式で取得
        today = datetime.datetime.now()
        base_date = today - datetime.timedelta(days=days_ago)
        yesterday = base_date - datetime.timedelta(days=1)
        today_str = base_date.strftime("%Y%m%d")
        yesterday_str = yesterday.strftime("%Y%m%d")
        
        # S3バケットとプレフィックス
        bucket_name = "adi-external-integration"
        prefix = "pallet-cloud/prod/"
        
        # 今日と昨日のファイルパス
        today_file = f"{today_str}_PC_commitment.csv"
        yesterday_file = f"{yesterday_str}_PC_commitment.csv"
        
        today_s3_path = f"{prefix}{today_file}"
        yesterday_s3_path = f"{prefix}{yesterday_file}"
        
        # S3クライアントの初期化
        s3_client = boto3.client('s3')
        
        # 一時ファイルパス
        temp_today_file = f"/tmp/{today_file}"
        temp_yesterday_file = f"/tmp/{yesterday_file}"
        
        # S3からファイルをダウンロード
        try:
            s3_client.download_file(bucket_name, today_s3_path, temp_today_file)
            print(f"Downloaded today's file from s3://{bucket_name}/{today_s3_path}")
        except Exception as e:
            print(f"Error downloading today's file: {e}")
            raise HTTPException(status_code=404, detail=f"Today's file not found: {today_file}")
            
        try:
            s3_client.download_file(bucket_name, yesterday_s3_path, temp_yesterday_file)
            print(f"Downloaded yesterday's file from s3://{bucket_name}/{yesterday_s3_path}")
        except Exception as e:
            print(f"Error downloading yesterday's file: {e}")
            raise HTTPException(status_code=404, detail=f"Yesterday's file not found: {yesterday_file}")
        
        # CSVファイルをDataFrameに読み込む（データクリーニング付き）
        def clean_dataframe(df):
            """DataFrameの文字列カラムをクリーニング"""
            for col in df.columns:
                if df[col].dtype == 'object':  # 文字列カラムの場合
                    # 空文字列をNaNに統一
                    df[col] = df[col].replace('', pd.NA)
                    # 先頭・末尾の空白を除去
                    df[col] = df[col].astype(str).str.strip()
                    # 'nan'文字列をNaNに変換
                    df[col] = df[col].replace('nan', pd.NA)
                    # 空白のみの文字列をNaNに変換
                    df[col] = df[col].replace(r'^\s*$', pd.NA, regex=True)
            return df
        
        today_df = pd.read_csv(temp_today_file, encoding='utf-8', keep_default_na=True, na_values=['', 'NULL', 'null', 'None'])
        yesterday_df = pd.read_csv(temp_yesterday_file, encoding='utf-8', keep_default_na=True, na_values=['', 'NULL', 'null', 'None'])
        
        # データクリーニング
        today_df = clean_dataframe(today_df)
        yesterday_df = clean_dataframe(yesterday_df)
        
        print(f"Today's file has {len(today_df)} rows")
        print(f"Yesterday's file has {len(yesterday_df)} rows")
        
        # 除外対象のcontract_idに一致する行を除外
        if len(today_df.columns) > 2 and exclude_contracts:
            contract_id_column = today_df.columns[2]  # 0-indexで2は3カラム目
            
            # デバッグ情報を出力
            print(f"Contract ID column: {contract_id_column}")
            print(f"Contract ID column dtype: {today_df[contract_id_column].dtype}")
            print(f"Sample contract IDs from today's file: {today_df[contract_id_column].head().tolist()}")
            print(f"Exclude contracts list (first 5): {exclude_contracts[:5]}")
            print(f"Exclude contracts types: {[type(x) for x in exclude_contracts[:5]]}")
            
            # 型変換: contract_idと除外リストを文字列に統一
            today_df[contract_id_column] = today_df[contract_id_column].astype(str)
            exclude_contracts_str = [str(x).strip() for x in exclude_contracts]
            
            # フィルタリング前の行数
            before_count = len(today_df)
            today_df = today_df[~today_df[contract_id_column].isin(exclude_contracts_str)]
            after_count = len(today_df)
            excluded_count = before_count - after_count
            
            print(f"Today's file: {before_count} -> {after_count} rows (excluded {excluded_count} rows)")
        
        if len(yesterday_df.columns) > 2 and exclude_contracts:
            contract_id_column = yesterday_df.columns[2]  # 0-indexで2は3カラム目
            
            # 型変換: contract_idと除外リストを文字列に統一
            yesterday_df[contract_id_column] = yesterday_df[contract_id_column].astype(str)
            exclude_contracts_str = [str(x).strip() for x in exclude_contracts]
            
            # フィルタリング前の行数
            before_count = len(yesterday_df)
            yesterday_df = yesterday_df[~yesterday_df[contract_id_column].isin(exclude_contracts_str)]
            after_count = len(yesterday_df)
            excluded_count = before_count - after_count
            
            print(f"Yesterday's file: {before_count} -> {after_count} rows (excluded {excluded_count} rows)")
        
        # 重複検知と除去機能を追加
        def detect_and_remove_duplicates(df, file_name):
            """重複を検知し、除去する"""
            if len(df.columns) == 0:
                return df, []
            
            primary_key = df.columns[0]
            duplicates = df[df.duplicated(subset=[primary_key], keep=False)]
            duplicate_keys = []
            
            if not duplicates.empty:
                duplicate_keys = duplicates[primary_key].unique().tolist()
                print(f"Found {len(duplicate_keys)} duplicate keys in {file_name}: {duplicate_keys}")
                # 重複を除去（最初の行を保持）
                df_deduped = df.drop_duplicates(subset=[primary_key], keep='first')
                print(f"After deduplication: {len(df_deduped)} rows (removed {len(df) - len(df_deduped)} duplicates)")
                return df_deduped, duplicate_keys
            else:
                print(f"No duplicates found in {file_name}")
                return df, duplicate_keys
        
        # 重複検知と除去を実行
        today_df, today_duplicates = detect_and_remove_duplicates(today_df, "today's file")
        yesterday_df, yesterday_duplicates = detect_and_remove_duplicates(yesterday_df, "yesterday's file")
        
        # 元のカラム順序を保存
        original_columns = today_df.columns.tolist()
        
        # 主キーを特定（1カラム目を主キーとする）
        if len(today_df.columns) > 1:
            primary_key = today_df.columns[0]
            
            # **最適化1: インデックスを設定して高速化**
            today_df_indexed = today_df.set_index(primary_key)
            yesterday_df_indexed = yesterday_df.set_index(primary_key)
            
            # **最適化2: 集合演算を使用**
            today_keys = set(today_df_indexed.index)
            yesterday_keys = set(yesterday_df_indexed.index)
            
            # 1. 今日のファイルにしかない行を抽出
            only_today_keys = today_keys - yesterday_keys
            only_in_today = today_df_indexed.loc[list(only_today_keys)] if only_today_keys else pd.DataFrame()
            
            # 2. 昨日のファイルにしかない行を抽出（削除された行）
            only_yesterday_keys = yesterday_keys - today_keys
            only_in_yesterday = yesterday_df_indexed.loc[list(only_yesterday_keys)] if only_yesterday_keys else pd.DataFrame()
            
            # 削除された行に2列目に1を挿入
            if not only_in_yesterday.empty and len(only_in_yesterday.columns) > 0:
                # 2列目の名前を取得（indexの次の列）
                second_column = only_in_yesterday.columns[0]
                # 2列目に1を設定
                only_in_yesterday[second_column] = 1
            
            # 3. 共通キーを取得
            common_keys = today_keys & yesterday_keys
            print(f"Processing {len(common_keys)} common keys for differences")
            
            # **最適化3: ベクトル化された比較**
            diff_details = []
            changed_indices = set()  # setを使用して重複排除を効率化
            detailed_diff_info = []  # デバッグ用の詳細情報
            
            if common_keys:
                # 共通キーのデータを一括で取得
                today_common = today_df_indexed.loc[list(common_keys)]
                yesterday_common = yesterday_df_indexed.loc[list(common_keys)]
                
                # **最適化4: カラムごとに一括比較（改良版）**
                for col in today_common.columns:
                    if col in yesterday_common.columns:
                        today_vals = today_common[col]
                        yesterday_vals = yesterday_common[col]
                        
                        # 数値型の場合は差分が1以上あるかチェック
                        if pd.api.types.is_numeric_dtype(today_vals):
                            # 両方がNaNでない場合の数値差分チェック
                            both_not_nan = pd.notna(today_vals) & pd.notna(yesterday_vals)
                            diff_mask = both_not_nan & (abs(today_vals - yesterday_vals) >= 1)
                            # 片方だけがNaNの場合もチェック
                            nan_diff_mask = today_vals.isna() != yesterday_vals.isna()
                            combined_mask = diff_mask | nan_diff_mask
                        else:
                            # 文字列型の場合：改良された比較ロジック
                            # NaNの扱いを統一（両方がNaNの場合は差分なし）
                            both_nan = pd.isna(today_vals) & pd.isna(yesterday_vals)
                            both_not_nan = pd.notna(today_vals) & pd.notna(yesterday_vals)
                            one_nan = pd.isna(today_vals) != pd.isna(yesterday_vals)
                            
                            # 両方が値を持つ場合の比較
                            value_diff = both_not_nan & (today_vals.astype(str) != yesterday_vals.astype(str))
                            
                            # 差分があるのは：値が異なる場合 OR 片方だけがNaNの場合
                            combined_mask = value_diff | one_nan
                        
                        # 差分があるキーをsetに追加（自動重複排除）
                        different_keys = combined_mask[combined_mask].index.tolist()
                        if different_keys:
                            changed_indices.update(different_keys)
                            
                            # デバッグ用：差分がある全件の詳細情報を記録
                            for key in different_keys:
                                today_val = today_vals.loc[key] if key in today_vals.index else 'KEY_NOT_FOUND'
                                yesterday_val = yesterday_vals.loc[key] if key in yesterday_vals.index else 'KEY_NOT_FOUND'
                                detailed_diff_info.append({
                                    'key': key,
                                    'column': col,
                                    'today_value': repr(today_val),
                                    'yesterday_value': repr(yesterday_val),
                                    'today_type': type(today_val).__name__,
                                    'yesterday_type': type(yesterday_val).__name__
                                })
                
                # setをリストに変換
                changed_indices = list(changed_indices)
                
                # 差分詳細記録
                for key in list(changed_indices):
                    diff_details.append({"key": key, "differences": ["Changes detected"]})
            
            # 変更された行を取得
            changed_rows = today_df_indexed.loc[changed_indices] if changed_indices else pd.DataFrame()
            
            # 差分ファイルを作成（今日にしかない行 + 変更された行 + 削除された行）
            diff_frames = []
            if not only_in_today.empty:
                # インデックスをリセットして元のカラム順序を復元
                only_in_today_reset = only_in_today.reset_index()
                only_in_today_ordered = only_in_today_reset[original_columns]
                diff_frames.append(only_in_today_ordered)
            
            if not changed_rows.empty:
                # インデックスをリセットして元のカラム順序を復元
                changed_rows_reset = changed_rows.reset_index()
                changed_rows_ordered = changed_rows_reset[original_columns]
                diff_frames.append(changed_rows_ordered)
            
            if not only_in_yesterday.empty:
                # インデックスをリセットして元のカラム順序を復元
                only_in_yesterday_reset = only_in_yesterday.reset_index()
                only_in_yesterday_ordered = only_in_yesterday_reset[original_columns]
                diff_frames.append(only_in_yesterday_ordered)
            
            if diff_frames:
                diff_df = pd.concat(diff_frames, ignore_index=True)
            else:
                diff_df = pd.DataFrame()
            
            # 差分ファイルをCSVに変換
            csv_buffer = io.StringIO()
            diff_df.to_csv(csv_buffer, index=False)
            
            # S3に出力（UTF-8エンコーディングを明示的に指定）
            output_s3_key = f"{prefix}output/commitment.csv"
            s3_client.put_object(
                Bucket=bucket_name,
                Key=output_s3_key,
                Body=csv_buffer.getvalue().encode('utf-8'),
                ContentType='text/csv; charset=utf-8'
            )
            
            # ローカルにも保存
            local_dir = f"./pallet_cloud/{today_str}"
            os.makedirs(local_dir, exist_ok=True)
            local_file_path = f"{local_dir}/commitment.csv"
            
            # CSVファイルを保存
            with open(local_file_path, 'w', encoding='utf-8') as f:
                f.write(csv_buffer.getvalue())
            
            # 差分詳細をテキストファイルに出力（デバッグ情報付き）
            diff_result_path = f"{local_dir}/commitment-diff-result.txt"
            with open(diff_result_path, 'w', encoding='utf-8') as f:
                f.write(f"Commitment Diff Results - {today_str} vs {yesterday_str}\n")
                f.write(f"Total rows in today's file: {len(today_df)}\n")
                f.write(f"Total rows in yesterday's file: {len(yesterday_df)}\n")
                
                # 除外情報を追加
                f.write(f"Excluded contracts: {len(exclude_contracts)} contracts excluded\n")
                
                # 重複情報を追加
                f.write(f"Duplicate keys in today's file: {len(today_duplicates)} - {today_duplicates}\n")
                f.write(f"Duplicate keys in yesterday's file: {len(yesterday_duplicates)} - {yesterday_duplicates}\n")
                
                f.write(f"New rows (only in today's file): {len(only_in_today)}\n")
                f.write(f"Deleted rows (only in yesterday's file): {len(only_in_yesterday)}\n")
                f.write(f"Changed rows: {len(changed_rows)}\n\n")
                
                # 除外詳細を出力
                if exclude_contracts:
                    f.write("=== CONTRACT EXCLUSION INFORMATION ===\n")
                    f.write(f"Excluded {len(exclude_contracts)} contract IDs from exclude_contracts.txt\n")
                    f.write("Target column: 3rd column (0-index: 2)\n")
                    f.write(f"Contract ID column name: {today_df.columns[2] if len(today_df.columns) > 2 else 'N/A'}\n")
                    f.write(f"Data type conversion: All values converted to string for comparison\n")
                    f.write(f"Sample exclude list: {exclude_contracts[:10]}\n\n")
                
                # 重複詳細を出力
                if today_duplicates or yesterday_duplicates:
                    f.write("=== DUPLICATE INFORMATION ===\n")
                    if today_duplicates:
                        f.write(f"Today's file had {len(today_duplicates)} duplicate keys: {today_duplicates}\n")
                    if yesterday_duplicates:
                        f.write(f"Yesterday's file had {len(yesterday_duplicates)} duplicate keys: {yesterday_duplicates}\n")
                    f.write("Note: Duplicates were automatically removed (keeping first occurrence)\n\n")
                
                if len(only_in_today) > 0:
                    f.write("=== NEW ROWS ===\n")
                    for key in only_in_today.index:
                        f.write(f"New row with key: {key}\n")
                    f.write("\n")
                
                if len(only_in_yesterday) > 0:
                    f.write("=== DELETED ROWS ===\n")
                    for key in only_in_yesterday.index:
                        f.write(f"Deleted row with key: {key}\n")
                    f.write("\n")
                
                if len(diff_details) > 0:
                    f.write("=== CHANGED ROWS ===\n")
                    for row_diff in diff_details:
                        f.write(f"Row with key '{row_diff['key']}' has differences\n")
                    f.write("\n")
                    
                    # デバッグ用の詳細情報を出力
                    if detailed_diff_info:
                        f.write(f"=== DETAILED DIFF INFO (All {len(detailed_diff_info)} differences) ===\n")
                        for info in detailed_diff_info:
                            f.write(f"Key: {info['key']}, Column: {info['column']}\n")
                            f.write(f"  Today: {info['today_value']} (type: {info['today_type']})\n")
                            f.write(f"  Yesterday: {info['yesterday_value']} (type: {info['yesterday_type']})\n")
                            f.write("\n")
                else:
                    f.write("No differences found in existing rows.\n")
            
            # 差分詳細ファイルをS3にもアップロード
            with open(diff_result_path, 'r', encoding='utf-8') as f:
                diff_detail_content = f.read()
            
            # S3に出力
            diff_detail_s3_key = f"{prefix}output/commitment-diff-result.txt"
            s3_client.put_object(
                Bucket=bucket_name,
                Key=diff_detail_s3_key,
                Body=diff_detail_content.encode('utf-8'),
                ContentType='text/plain; charset=utf-8'
            )
            
            print(f"Saved local file to: {local_file_path}")
            print(f"Saved diff details to: {diff_result_path}")
            print(f"Uploaded diff details to S3: s3://{bucket_name}/{diff_detail_s3_key}")
            
            output_s3_path = f"s3://{bucket_name}/{output_s3_key}"
            
            return {
                "message": "コミットメントファイルの差分計算が完了しました",
                "today_date": today_str,
                "yesterday_date": yesterday_str,
                "total_rows_today": len(today_df),
                "total_rows_yesterday": len(yesterday_df),
                "excluded_contracts_count": len(exclude_contracts),
                "today_duplicates": len(today_duplicates),
                "yesterday_duplicates": len(yesterday_duplicates),
                "duplicate_keys_today": today_duplicates,
                "duplicate_keys_yesterday": yesterday_duplicates,
                "new_rows": len(only_in_today),
                "deleted_rows": len(only_in_yesterday),
                "changed_rows": len(changed_indices),
                "diff_rows": len(diff_df),
                "output_file": output_s3_path,
                "debug_info": {
                    "detailed_diffs_found": len(detailed_diff_info),
                    "sample_differences": detailed_diff_info[:3] if detailed_diff_info else [],
                    "note": "All columns included in comparison (no excluded columns). Contract exclusion applied to 3rd column."
                }
            }
        else:
            raise HTTPException(status_code=400, detail="ファイルにカラムがありません")
            
    except (GoogleAPICallError, NotFound) as e:
        print(f"BigQuery Error: {e}")
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        print(f"Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    

@router.post('/gcp/pallet-cloud/compress')
def compress_pallet_cloud_files(params: DaysAgoParams = None):
    try:
        # パラメータがない場合はデフォルト値を使用
        days_ago = 0 if params is None else params.days_ago
        
        # 現在の日付と指定された日数前の日付をyyyymmdd形式で取得
        today = datetime.datetime.now()
        base_date = today - datetime.timedelta(days=days_ago)
        today_str = base_date.strftime("%Y%m%d")
        
        # ローカルディレクトリパス
        local_dir = f"./pallet_cloud/{today_str}"
        
        # S3クライアントの初期化
        s3_client = boto3.client('s3')
        bucket_name = "adi-external-integration"
        prefix = "pallet-cloud/prod/"
        
        # === STEP 1: CHECK用ZIP作成とS3アップロード ===
        
        # checkディレクトリ名とパス
        check_dir_name = f"{today_str}_check"
        check_dir_path = os.path.join(local_dir, check_dir_name)
        
        # checkディレクトリが既に存在する場合は削除
        if os.path.exists(check_dir_path):
            shutil.rmtree(check_dir_path)
        
        # checkディレクトリを作成
        os.makedirs(check_dir_path, exist_ok=True)
        print(f"Created check directory: {check_dir_path}")
        
        # ディレクトリ内の全ファイルをcheckディレクトリにコピー
        check_copied_files = []
        for item in os.listdir(local_dir):
            item_path = os.path.join(local_dir, item)
            # ディレクトリは除外
            if os.path.isfile(item_path):
                dest_path = os.path.join(check_dir_path, item)
                shutil.copy2(item_path, dest_path)
                check_copied_files.append(dest_path)
                print(f"Copied to check: {item} -> {check_dir_name}/")
        
        # ZIPファイル名とパス
        zip_filename = f"{today_str}_check.zip"
        zip_filepath = os.path.join(local_dir, zip_filename)
        
        print(f"Creating ZIP file: {zip_filepath}")
        
        # ZIPファイルを作成
        with zipfile.ZipFile(zip_filepath, 'w', zipfile.ZIP_DEFLATED) as zipf:
            # checkディレクトリ全体をZIPに追加
            for root, dirs, files in os.walk(check_dir_path):
                for file in files:
                    file_path = os.path.join(root, file)
                    # ZIP内でのパスを設定（check_dir_nameを含める）
                    arcname = os.path.relpath(file_path, local_dir)
                    zipf.write(file_path, arcname)
                    print(f"Added to ZIP: {arcname}")
        
        print(f"ZIP file created successfully: {zip_filepath}")
        print(f"ZIP file size: {os.path.getsize(zip_filepath)} bytes")
        
        # ZIPファイルをS3にアップロード
        zip_s3_key = f"{prefix}check/{zip_filename}"
        try:
            with open(zip_filepath, 'rb') as zip_file:
                s3_client.put_object(
                    Bucket=bucket_name,
                    Key=zip_s3_key,
                    Body=zip_file,
                    ContentType='application/zip'
                )
            print(f"Uploaded ZIP to S3: s3://{bucket_name}/{zip_s3_key}")
            zip_s3_path = f"s3://{bucket_name}/{zip_s3_key}"
        except Exception as e:
            print(f"Error uploading ZIP to S3: {e}")
            # S3アップロードに失敗してもメイン処理は続行
            zip_s3_path = None
        
        # checkディレクトリとZIPファイルをクリーンアップ
        shutil.rmtree(check_dir_path)
        print(f"Removed check directory: {check_dir_path}")
        os.remove(zip_filepath)
        print(f"Removed local ZIP file: {zip_filepath}")
        
        # === STEP 2: 通常のTAR.GZ作成処理 ===
        
        # 必要なファイルリスト
        required_files = [
            "building.csv",
            "room.csv", 
            "contract2.csv",
            "tenant.csv",
            "contract_tenant.csv",
            "contract_resident.csv",
            "commitment.csv"
        ]
        
        # ファイルの存在確認
        missing_files = []
        existing_files = []
        
        for file_name in required_files:
            file_path = os.path.join(local_dir, file_name)
            if os.path.exists(file_path):
                existing_files.append(file_path)
                print(f"✓ Found: {file_name}")
            else:
                missing_files.append(file_name)
                print(f"✗ Missing: {file_name}")
        
        # 全てのファイルが存在するかチェック
        if missing_files:
            raise HTTPException(
                status_code=404, 
                detail=f"Missing required files: {', '.join(missing_files)}"
            )
        
        # 一時ディレクトリを作成
        temp_dir_name = f"mdi_palettecloud_{today_str}"
        temp_dir_path = os.path.join(local_dir, temp_dir_name)
        
        # 一時ディレクトリが既に存在する場合は削除
        if os.path.exists(temp_dir_path):
            shutil.rmtree(temp_dir_path)
        
        # 一時ディレクトリを作成
        os.makedirs(temp_dir_path, exist_ok=True)
        print(f"Created temporary directory: {temp_dir_path}")
        
        # ファイルを一時ディレクトリにコピー
        copied_files = []
        for file_path in existing_files:
            file_name = os.path.basename(file_path)
            dest_path = os.path.join(temp_dir_path, file_name)
            shutil.copy2(file_path, dest_path)
            copied_files.append(dest_path)
            print(f"Copied: {file_name} -> {temp_dir_name}/")
        
        # tarファイル名
        tar_filename = f"mdi_palettecloud_{today_str}.tar"
        tar_filepath = os.path.join(local_dir, tar_filename)
        
        # gzipファイル名
        gzip_filename = f"mdi_palettecloud_{today_str}.tar.gz"
        gzip_filepath = os.path.join(local_dir, gzip_filename)
        
        print(f"Creating tar file: {tar_filepath}")
        
        # tarファイルを作成（ディレクトリ構造を含む）
        with tarfile.open(tar_filepath, 'w') as tar:
            # local_dirを基準として相対パスでディレクトリ全体を追加
            tar.add(temp_dir_path, arcname=temp_dir_name)
            print(f"Added directory to tar: {temp_dir_name}/")
        
        print(f"Tar file created successfully: {tar_filepath}")
        print(f"Tar file size: {os.path.getsize(tar_filepath)} bytes")
        
        # gzipで圧縮
        print(f"Compressing to gzip: {gzip_filepath}")
        
        with open(tar_filepath, 'rb') as f_in:
            with gzip.open(gzip_filepath, 'wb') as f_out:
                shutil.copyfileobj(f_in, f_out)
        
        print(f"Gzip file created successfully: {gzip_filepath}")
        print(f"Gzip file size: {os.path.getsize(gzip_filepath)} bytes")
        
        # 一時ディレクトリを削除
        shutil.rmtree(temp_dir_path)
        print(f"Removed temporary directory: {temp_dir_path}")
        
        # 元のtarファイルを削除（gzipファイルのみ残す）
        os.remove(tar_filepath)
        print(f"Removed original tar file: {tar_filepath}")
        
        # 圧縮率を計算
        original_total_size = sum(os.path.getsize(fp) for fp in existing_files)
        compressed_size = os.path.getsize(gzip_filepath)
        compression_ratio = (1 - compressed_size / original_total_size) * 100
        
        return {
            "message": "ファイルの圧縮が完了しました",
            "date": today_str,
            "compressed_file": gzip_filepath,
            "files_included": [os.path.basename(fp) for fp in existing_files],
            "file_count": len(existing_files),
            "original_total_size": original_total_size,
            "compressed_size": compressed_size,
            "compression_ratio": f"{compression_ratio:.1f}%",
            "compression_details": {
                "directory_name": temp_dir_name,
                "tar_filename": tar_filename,
                "gzip_filename": gzip_filename,
                "output_directory": local_dir
            },
            "check_zip_info": {
                "zip_filename": zip_filename,
                "check_files_count": len(check_copied_files),
                "s3_upload_path": zip_s3_path,
                "s3_upload_success": zip_s3_path is not None
            }
        }
        
    except FileNotFoundError as e:
        print(f"Directory not found: {e}")
        raise HTTPException(status_code=404, detail=f"Directory not found: {local_dir}")
    except PermissionError as e:
        print(f"Permission error: {e}")
        raise HTTPException(status_code=403, detail=f"Permission error: {str(e)}")
    except Exception as e:
        print(f"Error during compression: {e}")
        raise HTTPException(status_code=500, detail=f"Compression error: {str(e)}")